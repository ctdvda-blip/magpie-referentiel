# -*- coding: utf-8 -*-
"""Produit l'indice B : un referentiel UNIFIE, sans trace de la partition
entre les 19 notions du tableau d'origine et les 97 issues des programmes.

Par rapport a l'indice A :
    - onglet "Fondamentaux V1" supprime
    - colonnes de provenance et de couverture supprimees :
      Origine / couverture, D3, P3, P6, P6b, P6A, RG8, Nb prog., Ordre V1, Ref. croisee
    - identifiants unifies REF-001..REF-116, sans prefixe distinguant l'origine
    - colonne Notes nettoyee de toute mention de provenance
    - onglet Synthese recalcule, sans le bloc d'ecarts V1 / programmes

L'ordre des lignes de l'indice A est conserve tel quel : il porte le classement
par domaine puis par categorie, et la progression pedagogique.
"""
import os
import sys
import io
import csv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ICI = os.path.dirname(os.path.abspath(__file__))
if ICI not in sys.path:
    sys.path.insert(0, ICI)

PROJET = os.path.abspath(os.path.join(ICI, "..", ".."))
SOURCE = os.path.join(PROJET, "Fondamentaux Grasshopper - IndA - 25-08-2026.xlsx")
SORTIE = os.path.join(PROJET, "Fondamentaux Grasshopper - IndB - 26-08-2026.xlsx")
EXPORTS = os.path.join(PROJET, "EXPORTS")
ARCHIVE = os.path.join(PROJET, "Anciens fichiers")

VERSION = "Ind. B"
DATE_FR = "26/08/2026"

# colonnes de l'indice A, par indice 0
GARDEES = [0, 1, 2, 3, 4, 5, 6, 17, 18, 19, 20, 21, 22, 23, 24]
ENTETES = [u"N°", u"ID", u"Domaine", u"Catégorie", u"Notion", u"Description",
           u"Niveau", u"ValidationMode suggéré", u"Type d'exercice",
           u"Priorité de production", u"Nb exercices Magpie prévus",
           u"À réaliser par", u"Statut", u"Lien", u"Notes",
           u"Nature pédagogique", u"Exercices Magpie"]
LARGEURS = [5, 10, 34, 32, 40, 70, 15, 22, 24, 12, 13, 18, 14, 26, 60, 22, 26]

# ---------------------------------------------------------------------------
# Nature pedagogique de chaque notion.
#
# La skill de conception avertit que les referentiels melangent couramment
# connaissances et competences, et demande de le DIRE : c'est une information
# utile au formateur, pas une critique du referentiel.
#
# La nature n'est pas devinee a partir du libelle de la notion — ce serait une
# lecture de surface. Elle est deduite de la maniere dont les exercices Magpie
# la traitent reellement :
#   - « Compétence »   : au moins un exercice note la mobilise ;
#   - « Connaissance » : elle n'est traitee que par une question charniere ;
#   - « À qualifier »  : aucun exercice ne la couvre encore.
# ---------------------------------------------------------------------------


def nature_des_notions():
    """Retourne {ID de notion: (nature, exercices)} d'apres le lot A produit."""
    import re as _re
    try:
        from lots import TOUS as corpus
    except Exception:
        return {}
    par_ref = {}
    for ex in corpus:
        for rid in _re.findall(u"REF-[0-9]+", ex.get("ref", u"")):
            par_ref.setdefault(rid, []).append(
                (ex["id"], ex.get("verdict") or u"competence"))
    out = {}
    for rid, items in par_ref.items():
        ids = sorted(i for i, _v in items)
        if any(v == u"competence" for _i, v in items):
            nat = u"Compétence"
        else:
            nat = u"Connaissance"
        out[rid] = (nat, u", ".join(ids))
    return out

# note factuelle conservee, reformulee sans reference a l'origine de la notion
ALERTE_MATCHING = (
    u"⚠ Vérifié dans Rhino 8 le 26/08/2026 : la correspondance PAR DÉFAUT de "
    u"Grasshopper n'est pas la troncature sur la liste la plus courte, mais la "
    u"correspondance sur la liste la plus LONGUE, la liste courte étant prolongée "
    u"par répétition de son dernier élément. Une liste de 10 et une liste de 4 dans "
    u"une Addition produisent 10 résultats, pas 4. Le mode Shortest List existe mais "
    u"doit être demandé explicitement par clic droit. L'intitulé de cette notion est "
    u"donc à corriger."
)

NOIR = "1F2A37"
BLEU = "2E5C8A"
# La numerotation des domaines commence a 1, et non a 0 : un domaine « 0 »
# se lit comme un hors-serie alors que le socle Rhino est bien la premiere
# etape du parcours. RENUM porte l'ancien libelle vers le nouveau ; les
# identifiants REF- ne changent pas, la tracabilite est donc preservee.
RENUM = {
    u"0 – Socle Rhino (prérequis)": u"1 – Socle Rhino (prérequis)",
    u"1 – Environnement et principes Grasshopper":
        u"2 – Environnement et principes Grasshopper",
    u"2 – Données et logique": u"3 – Données et logique",
    u"3 – Géométrie paramétrique": u"4 – Géométrie paramétrique",
    u"4 – Mesures, quantitatifs et export": u"5 – Mesures, quantitatifs et export",
    u"5 – Méthode, performance et évènements":
        u"6 – Méthode, performance et évènements",
    u"6 – Algorithmique avancée": u"7 – Algorithmique avancée",
    u"7 – Développement, scripting et API": u"8 – Développement, scripting et API",
    u"8 – Interfaces, web et interopérabilité":
        u"9 – Interfaces, web et interopérabilité",
    u"9 – Aide à la fabrication": u"10 – Aide à la fabrication",
}

ACC = {u"1 – Socle Rhino (prérequis)": "E8E8E8",
       u"2 – Environnement et principes Grasshopper": "DDE8F5",
       u"3 – Données et logique": "D9EAD3",
       u"4 – Géométrie paramétrique": "FCE5CD",
       u"5 – Mesures, quantitatifs et export": "FFF2CC",
       u"6 – Méthode, performance et évènements": "EAD1DC",
       u"7 – Algorithmique avancée": "D9D2E9",
       u"8 – Développement, scripting et API": "D0E0E3",
       u"9 – Interfaces, web et interopérabilité": "F4CCCC",
       u"10 – Aide à la fabrication": "E6E6C8",
       u"11 – IA et assistance générative": "E3D9F5"}

THIN = Side(style="thin", color="BFC7D1")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HFONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Calibri", size=10)
TFONT = Font(name="Calibri", size=16, bold=True, color=NOIR)
VFONT = Font(name="Calibri", size=10, bold=True, color=BLEU)


def entete(ws, titre, ncol):
    ws.cell(row=1, column=1, value=titre).font = TFONT
    c = ws.cell(row=1, column=max(ncol, 2),
                value=u"Fondamentaux Grasshopper – %s – %s" % (VERSION, DATE_FR))
    c.font = VFONT
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 24


def table_header(ws, row, headers, widths):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HFONT
        c.fill = PatternFill("solid", fgColor=BLEU)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORD
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[row].height = 32


def nettoyer_note(note, notion):
    """Retire toute mention de provenance ; conserve l'avertissement factuel."""
    n = (note or u"").strip()
    if u"Data Matching par défaut" in (notion or u""):
        return ALERTE_MATCHING
    if n.startswith(u"Notion issue du tableau des fondamentaux"):
        return u""
    return n


def main():
    src = openpyxl.load_workbook(SOURCE, data_only=False)
    ref = src[u"Référentiel"]

    lignes, mapping = [], []
    for r in ref.iter_rows(min_row=4, max_row=ref.max_row, values_only=True):
        if r[0] is None:
            continue
        ancien = r[1]
        nouveau = u"REF-%03d" % (len(lignes) + 1)
        mapping.append((ancien, nouveau))
        l = [r[i] if i < len(r) else None for i in GARDEES]
        l[0] = len(lignes) + 1
        l[1] = nouveau
        l[14] = nettoyer_note(l[14], l[4])
        if l[2] in RENUM:
            l[2] = RENUM[l[2]]
        lignes.append([u"" if v is None else v for v in l])
    print(u"Lignes reprises : %d" % len(lignes))

    # ------------------------------------------------ domaine 11 : IA
    try:
        import domaine_ia
        ajout = domaine_ia.lignes_referentiel(len(lignes) + 1)
        lignes.extend(ajout)
        print(u"Domaine IA ajouté : %d notions (REF-%03d à REF-%03d)"
              % (len(ajout), len(lignes) - len(ajout) + 1, len(lignes)))
    except Exception as _ex:
        print(u"ATTENTION : domaine IA non ajouté (%s)" % _ex)

    corr = dict(mapping)
    out = openpyxl.Workbook()

    # ---------------------------------------------------------- Lisez-moi
    ws = out.active
    ws.title = u"Lisez-moi"
    entete(ws, u"Fondamentaux Grasshopper – Référentiel unifié des notions Magpie", 4)
    for col, w in zip("ABCD", (34, 95, 18, 40)):
        ws.column_dimensions[col].width = w
    bloc = [
        (u"", u""),
        (u"Projet", u"MAGPIE – outil d'exercices Grasshopper autocorrigés (RhinoForYou)"),
        (u"Objet du document", u"Référentiel unique des notions Rhino / Grasshopper à couvrir "
                               u"par les exercices Magpie."),
        (u"Indice", VERSION),
        (u"Date", DATE_FR),
        (u"Établi par", u"Charles THIERRY DE VILLE D'AVRAY"),
        (u"", u""),
        (u"CE QUI CHANGE PAR RAPPORT À L'INDICE A", u""),
        (u"   Référentiel unifié", u"Les %d notions forment désormais un ensemble homogène. "
                                   u"Plus aucune colonne, ni aucun identifiant, ne distingue "
                                   u"les notions selon leur provenance." % len(lignes)),
        (u"   Identifiants", u"REF-001 à REF-%03d, attribués dans l'ordre de lecture." % len(lignes)),
        (u"   Onglet supprimé", u"« Fondamentaux V1 » : son contenu est fondu dans le référentiel."),
        (u"   Colonnes supprimées", u"Origine / couverture, D3, P3, P6, P6b, P6A, RG8, "
                                    u"Nb prog., Ordre V1, Réf. croisée."),
        (u"", u""),
        (u"CONTENU DU CLASSEUR", u""),
        (u"   Référentiel", u"%d notions, classées par domaine puis par catégorie." % len(lignes)),
        (u"   Contenu programmes", u"Relevé verbatim des items des programmes de formation, "
                                   u"rattachés à une ligne du référentiel. Conservé comme "
                                   u"pièce justificative des sources."),
        (u"   Synthèse", u"Comptages par domaine, niveau, mode de validation et type d'exercice."),
        (u"   Listes", u"Listes de valeurs des menus déroulants du référentiel."),
        (u"", u""),
        (u"LÉGENDE — MODE DE VALIDATION", u""),
        (u"   ExactOrderedList", u"Ordre des éléments significatif."),
        (u"   SetEquality", u"Ordre indifférent, ensemble attendu."),
        (u"   SingleValue", u"Une valeur unique attendue."),
        (u"   NumericTolerance", u"Valeur numérique avec marge d'erreur."),
        (u"   GeometryTolerance", u"Comparaison géométrique avec tolérance."),
        (u"   Conceptuel (QCM)", u"Notion non vérifiable par comparaison de résultat."),
        (u"", u""),
        (u"POINT À ARBITRER", u""),
        (u"   Correspondance par défaut", u"L'intitulé de la notion « Data Matching par défaut "
                                          u"(shortest list) » est inexact : voir l'avertissement "
                                          u"porté en colonne Notes de cette ligne."),
        (u"", u""),
        (u"JOURNAL DES INDICES", u""),
    ]
    for a, b in bloc:
        ws.append([a, b])
    r0 = ws.max_row + 1
    for j, t in enumerate((u"Indice", u"Objet", u"Date", u"Auteur"), 1):
        c = ws.cell(row=r0, column=j, value=t)
        c.font = HFONT
        c.fill = PatternFill("solid", fgColor=BLEU)
        c.border = BORD
    ws.cell(row=r0 + 1, column=1, value=u"Ind. A")
    ws.cell(row=r0 + 1, column=2, value=u"Création. Fusion du tableau d'origine et du contenu "
                                        u"des 6 programmes de formation, avec matrice de "
                                        u"couverture et analyse des écarts.")
    ws.cell(row=r0 + 1, column=3, value=u"25/08/2026")
    ws.cell(row=r0 + 1, column=4, value=u"C. THIERRY DE VILLE D'AVRAY")
    ws.cell(row=r0 + 2, column=1, value=u"Ind. B")
    ws.cell(row=r0 + 2, column=2, value=u"Unification. Suppression de l'onglet « Fondamentaux V1 », "
                                        u"des colonnes de provenance et de couverture, et des "
                                        u"préfixes d'identifiant. Identifiants renumérotés "
                                        u"REF-001 à REF-%03d." % len(lignes))
    ws.cell(row=r0 + 2, column=3, value=DATE_FR)
    ws.cell(row=r0 + 2, column=4, value=u"C. THIERRY DE VILLE D'AVRAY")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        for c in row:
            if not c.font.bold:
                c.font = BFONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
    for rr in range(2, ws.max_row + 1):
        v = ws.cell(row=rr, column=1).value
        if v and not v.startswith(u" ") and not ws.cell(row=rr, column=2).value:
            ws.cell(row=rr, column=1).font = Font(name="Calibri", size=10,
                                                  bold=True, color=BLEU)
    ws.sheet_view.showGridLines = False

    # -------------------------------------------------------- Référentiel
    ws = out.create_sheet(u"Référentiel")
    entete(ws, u"Référentiel des notions – Rhino / Grasshopper", len(ENTETES))
    table_header(ws, 3, ENTETES, LARGEURS)
    _nat = nature_des_notions()
    for l in lignes:
        n, ex = _nat.get(l[1], (u"À qualifier", u""))
        ws.append(list(l) + [n, ex])
    dernier = ws.max_row
    for i, l in enumerate(lignes):
        rr = 4 + i
        fill = PatternFill("solid", fgColor=ACC.get(l[2], "FFFFFF"))
        for j in range(1, len(ENTETES) + 1):
            c = ws.cell(row=rr, column=j)
            c.font = BFONT
            c.border = BORD
            c.alignment = Alignment(vertical="top",
                                    wrap_text=j in (3, 4, 5, 6, 15, 17),
                                    horizontal="center" if j in (1, 7, 10, 11, 13, 16) else "left")
            if j == 16:
                _v = c.value
                if _v == u"Connaissance":
                    c.font = Font(name="Calibri", size=10, bold=True, color="B06A00")
                elif _v == u"À qualifier":
                    c.font = Font(name="Calibri", size=10, color="8A8A8A")
                elif _v == u"Compétence":
                    c.font = Font(name="Calibri", size=10, bold=True, color="1F7A3D")
            if j in (3, 4):
                c.fill = fill
        if l[14]:
            ws.cell(row=rr, column=15).font = Font(name="Calibri", size=10, color="B00020")
    ws.freeze_panes = "F4"
    ws.auto_filter.ref = "A3:%s%d" % (get_column_letter(len(ENTETES)), dernier)
    ws.sheet_view.showGridLines = False

    # ------------------------------------------------- Contenu programmes
    scp = src[u"Contenu programmes"]
    ws = out.create_sheet(u"Contenu programmes")
    entete(ws, u"Relevé exhaustif du contenu des programmes de formation (verbatim)", 10)
    H = [c.value for c in scp[3]]
    table_header(ws, 3, H, [5, 7, 34, 10, 14, 34, 78, 12, 32, 30])
    n = 0
    for r in scp.iter_rows(min_row=4, max_row=scp.max_row, values_only=True):
        if r[0] is None:
            continue
        l = list(r)
        l[7] = corr.get(l[7], l[7])
        ws.append(l)
        n += 1
    for rr in range(4, ws.max_row + 1):
        for j in range(1, 11):
            c = ws.cell(row=rr, column=j)
            c.font = BFONT
            c.border = BORD
            c.alignment = Alignment(vertical="top", wrap_text=j in (3, 6, 7, 9, 10),
                                    horizontal="center" if j in (1, 2, 8) else "left")
    ws.freeze_panes = "D4"
    ws.auto_filter.ref = "A3:J%d" % ws.max_row
    ws.sheet_view.showGridLines = False
    print(u"Contenu programmes : %d lignes, identifiants remappés" % n)

    # ------------------------------------------------------------ Synthèse
    ws = out.create_sheet(u"Synthèse")
    entete(ws, u"Synthèse", 3)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    r = 3

    def bloc_compte(titre, col_ref, valeurs, r):
        ws.cell(row=r, column=1, value=titre).font = Font(name="Calibri", size=11,
                                                          bold=True, color=BLEU)
        r += 1
        for j, h in enumerate((titre.split(u" par ")[-1].capitalize(), u"Nb notions"), 1):
            c = ws.cell(row=r, column=j, value=h)
            c.font = HFONT
            c.fill = PatternFill("solid", fgColor=BLEU)
            c.border = BORD
            c.alignment = Alignment(horizontal="center")
        r += 1
        for v in valeurs:
            ws.cell(row=r, column=1, value=v).font = BFONT
            ws.cell(row=r, column=1).border = BORD
            c = ws.cell(row=r, column=2,
                        value=u'=COUNTIF(Référentiel!$%s$4:$%s$%d,$A$%d)'
                              % (col_ref, col_ref, dernier, r))
            c.font = BFONT
            c.border = BORD
            c.alignment = Alignment(horizontal="center")
            r += 1
        return r + 1

    ordre = [k for k in sorted(ACC)]
    r = bloc_compte(u"Répartition par domaine", "C", ordre, r)
    r = bloc_compte(u"Répartition par niveau", "G",
                    [u"Débutant", u"Intermédiaire", u"Perfectionnement", u"Expert"], r)
    r = bloc_compte(u"Répartition par mode de validation", "H",
                    [u"ExactOrderedList", u"SetEquality", u"SingleValue",
                     u"NumericTolerance", u"GeometryTolerance", u"Conceptuel (QCM)"], r)
    r = bloc_compte(u"Répartition par type d'exercice", "I",
                    [u"Exercice Grasshopper", u"Exercice de synthèse",
                     u"QCM / question ciblée", u"Démonstration formateur"], r)
    ws.cell(row=r, column=1, value=u"Total des notions").font = Font(name="Calibri",
                                                                     size=11, bold=True)
    ws.cell(row=r, column=2, value=len(lignes)).font = Font(name="Calibri", size=11, bold=True)
    ws.sheet_view.showGridLines = False

    # -------------------------------------------------------------- Listes
    sl = src[u"Listes"]
    ws = out.create_sheet(u"Listes")
    entete(ws, u"Listes de valeurs", 6)
    for j in range(1, 7):
        ws.column_dimensions[get_column_letter(j)].width = 34
        for i in range(3, sl.max_row + 1):
            v = sl.cell(row=i, column=j).value
            if v is None:
                continue
            c = ws.cell(row=i, column=j, value=v)
            if i == 3:
                c.font = HFONT
                c.fill = PatternFill("solid", fgColor=BLEU)
                c.alignment = Alignment(horizontal="center", wrap_text=True)
            else:
                c.font = BFONT
            c.border = BORD
    ws.sheet_view.showGridLines = False

    try:
        out.save(SORTIE)
        print(u"écrit : %s" % SORTIE)
    except PermissionError:
        print(u"NON ÉCRIT : le classeur est ouvert dans Excel.")
        print(u"  %s" % SORTIE)
        print(u"  Fermez-le puis relancez ce script. Aucun autre classeur "
              u"ouvert n'a été touché.")

    # ------------------------------------------------- table de correspondance
    if not os.path.isdir(EXPORTS):
        os.makedirs(EXPORTS)
    chemin = os.path.join(EXPORTS, u"Correspondance_identifiants_IndA_vers_IndB.csv")
    fh = io.open(chemin, "w", encoding="utf-8-sig", newline="")
    w = csv.writer(fh, delimiter=str(";"), lineterminator="\n")
    w.writerow([u"Identifiant Ind. A", u"Identifiant Ind. B"])
    for a, b in mapping:
        w.writerow([a, b])
    fh.close()
    print(u"écrit : %s" % chemin)


if __name__ == "__main__":
    main()
