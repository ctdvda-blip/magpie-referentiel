# -*- coding: utf-8 -*-
"""Genere 'Fondamentaux Grasshopper - IndA - 25-08-2026.xlsx' pour le projet MAGPIE."""
import os, re, sys, zipfile, unicodedata
from xml.etree import ElementTree as ET
from collections import OrderedDict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from meta import (META, FOND_MAP, DOMAINES, NIVEAUX, MODES, TYPES, CAT_ORDER_D2, ALERTES,
                  D0, D1, D2, D3, D4, D5, D6, D7, D8, D9,
                  DEB, INT, PER, EXP)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJ = r"C:\Users\charl\.claude\projects\MAGPIE"
SRC = os.path.join(PROJ, "EXEMPLES PROGRAMMES DE FORMATION")
VERSION = "Ind. A"
DATE_FR = "25/08/2026"
OUT = os.path.join(PROJ, "Fondamentaux Grasshopper - IndA - 25-08-2026.xlsx")

W = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'

PROGS = [
    ('D3',  'Grasshopper débutant 3 jours',
     'Programme de formation Grasshopper d\u00e9butant 3 jours.docx'),
    ('P3',  'Grasshopper perfectionnement 3 jours',
     'Programme de formation Grasshopper perfectionnement 3 jours.docx'),
    ('P6',  'Grasshopper perfectionnement 6 jours',
     'Programme de formation Grasshopper perfectionnement 6 jours.docx'),
    ('P6b', 'Grasshopper perfectionnement 6 jours (1)',
     'Programme de formation Grasshopper perfectionnement 6 jours (1).docx'),
    ('P6A', 'Grasshopper perfectionnement 6 jours – Ind A – 27/02/2025',
     'Programme de formation Grasshopper perfectionnement 6 jours - IndA - 27-02-2025.docx'),
    ('RG8', 'Rhino – Grasshopper 8 jours',
     'Programme de formation Grasshopper 8 jours.docx'),
]
PCODES = [p[0] for p in PROGS]


# ----------------------------------------------------------------- extraction
def ptext(p):
    parts = []
    for n in p.iter():
        if n.tag == W + 't':
            parts.append(n.text or '')
        elif n.tag == W + 'tab':
            parts.append(' ')
    return re.sub(r'\s+', ' ', ''.join(parts)).strip()


def pstyle(p):
    pPr = p.find(W + 'pPr')
    s = ''
    if pPr is not None:
        ps = pPr.find(W + 'pStyle')
        if ps is not None:
            s = ps.get(W + 'val') or ''
        if pPr.find(W + 'numPr') is not None:
            s += '|LIST'
    return s


def norm(t):
    t = unicodedata.normalize('NFKD', t.lower())
    t = ''.join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r'[^a-z0-9]+', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()


def parse_programmes():
    rows = []
    for code, label, fn in PROGS:
        z = zipfile.ZipFile(os.path.join(SRC, fn))
        root = ET.fromstring(z.read('word/document.xml'))
        jour = session = rub = ''
        skipped_title = False
        for p in root.iter(W + 'p'):
            t = ptext(p)
            if not t:
                continue
            s = pstyle(p)
            if not skipped_title and 'Programme de formation' in t:
                skipped_title = True
                continue
            if re.match(r'^Jour\s*\d', t, re.I):
                jour, session, rub = t, '', ''
                continue
            if re.match(r'^Session', t, re.I):
                session, rub = t, ''
                continue
            if 'LIST' not in s:
                rub = t
                continue
            if code == 'P3' and t == t.upper() and len(t) > 3:
                rub = t.capitalize()
                continue
            rows.append(dict(prog=code, prog_label=label, jour=jour, session=session,
                             rubrique=rub, item=t, norm=norm(t)))
    return rows


def short_loc(jour, session):
    j = re.sub(r'^Jour\s*', 'J', jour, flags=re.I)
    if session:
        s = re.sub(r'^Session\s*N?\s*°?\s*', 'S', session, flags=re.I)
        return u'%s\u00b7%s' % (j, s)
    return j


# ------------------------------------------------------------- fondamentaux V1
def lire_fondamentaux_v1():
    wb = openpyxl.load_workbook(os.path.join(PROJ, "Fondamentaux Grasshopper.xlsx"))
    ws = wb[wb.sheetnames[0]]
    out = []
    for r in ws.iter_rows(min_row=2, max_row=20, max_col=8, values_only=True):
        if r[0] is None:
            continue
        out.append(dict(ordre=float(r[0]), cat=r[1], notion=r[2], desc=r[3],
                        mode=r[4], qui=r[5] or '', statut=r[6] or '', lien=r[7] or ''))
    return out


# ------------------------------------------------------------------ construction
rows = parse_programmes()
items = OrderedDict()
for r in rows:
    items.setdefault(r['norm'], []).append(r)
assert len(items) == 97, len(items)

fonds = lire_fondamentaux_v1()
assert len(fonds) == 19, len(fonds)


PRIO = {DEB: "P0", INT: "P1", PER: "P2", EXP: "P2"}

# --- lignes du referentiel
ref = []
for idx, (key, occ) in enumerate(items.items(), 1):
    dom, cat, notion, niv, mode, typ = META[idx]
    cov = {}
    for o in occ:
        cov[o['prog']] = short_loc(o['jour'], o['session'])
    ref.append(dict(
        src='PRG', idx=idx, dom=dom, cat=cat, notion=notion,
        desc=occ[0]['item'], rubrique=occ[0]['rubrique'], niv=niv, mode=mode,
        typ=typ, cov=cov, nbprog=len(cov), ordre_v1='', reff='',
        origine="Programme de formation", qui='', statut='', lien='', notes=''))

COUV_LABEL = {
    "Couverte": "Fondamentaux V1 – couverte par les programmes",
    "Partielle": "Fondamentaux V1 – couverture partielle",
    "Non couverte": "Fondamentaux V1 – hors programmes (à ajouter)",
}
for f in fonds:
    o = int(f['ordre'])
    couv, maps, dom, cat, niv = FOND_MAP[o]
    refs = ', '.join('PRG-%03d' % int(x) for x in maps.split(',') if x.strip())
    ref.append(dict(
        src='FND', idx=o, dom=dom, cat=cat, notion=f['notion'], desc=f['desc'],
        rubrique=f['cat'], niv=niv, mode=f['mode'], typ="Exercice Grasshopper",
        cov={}, nbprog=0, ordre_v1=f['ordre'], reff=refs,
        origine=COUV_LABEL[couv], qui=f['qui'], statut=f['statut'], lien=f['lien'],
        notes="Notion issue du tableau des fondamentaux Magpie V1." +
              (" Rapprochement programme : " + refs if refs else
               " Aucun item de programme correspondant.") +
              ((" " + ALERTES[o]) if o in ALERTES else "")))

# ordre des categories : D2 impose, sinon ordre de premiere apparition
cat_seq = {}
for c in CAT_ORDER_D2:
    cat_seq[(D2, c)] = len(cat_seq)
for r in ref:
    k = (r['dom'], r['cat'])
    if k not in cat_seq:
        cat_seq[k] = 1000 + len(cat_seq)

dom_seq = {d: i for i, d in enumerate(DOMAINES)}
ref.sort(key=lambda r: (dom_seq[r['dom']], cat_seq[(r['dom'], r['cat'])],
                        0 if r['src'] == 'FND' else 1, r['idx']))

for n, r in enumerate(ref, 1):
    r['ordre'] = n
    r['id'] = ('FND-%02d' % r['idx']) if r['src'] == 'FND' else ('PRG-%03d' % r['idx'])

id_by_idx = {r['idx']: r['id'] for r in ref if r['src'] == 'PRG'}

# ------------------------------------------------------------------ mise en forme
NOIR = "1F2A37"
BLEU = "2E5C8A"
GRIS = "EDF1F5"
ACC = {D0: "E8E8E8", D1: "DDE8F5", D2: "D9EAD3", D3: "FCE5CD", D4: "FFF2CC",
       D5: "EAD1DC", D6: "D9D2E9", D7: "D0E0E3", D8: "F4CCCC", D9: "E6E6C8"}

thin = Side(style='thin', color="BFC7D1")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
HFONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Calibri", size=10)
TFONT = Font(name="Calibri", size=16, bold=True, color=NOIR)
VFONT = Font(name="Calibri", size=10, bold=True, color=BLEU)

wb = openpyxl.Workbook()


def entete(ws, titre, ncol):
    """Titre + version discrete en haut a droite (regle d'affichage de version)."""
    ws.cell(row=1, column=1, value=titre).font = TFONT
    c = ws.cell(row=1, column=max(ncol, 2), value="Fondamentaux Grasshopper – %s – %s" % (VERSION, DATE_FR))
    c.font = VFONT
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 24


def table_header(ws, row, headers, widths):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HFONT
        c.fill = PatternFill("solid", fgColor=BLEU)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORD
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[row].height = 34


# ============================================================ 1. LISEZ-MOI
ws = wb.active
ws.title = "Lisez-moi"
entete(ws, "Fondamentaux Grasshopper – Référentiel des notions Magpie", 4)
ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 95
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 40

lignes = [
    ("", ""),
    ("Projet", "MAGPIE – outil d'exercices Grasshopper autocorrigés (RhinoForYou)"),
    ("Objet du document", "Référentiel unique des notions Rhino / Grasshopper à couvrir par les exercices Magpie, "
                          "obtenu en fusionnant le tableau des fondamentaux V1 et l'intégralité du contenu des "
                          "programmes de formation du catalogue."),
    ("Indice", VERSION),
    ("Date", DATE_FR),
    ("Établi par", "Charles THIERRY DE VILLE D'AVRAY"),
    ("Document source remplacé", "Fondamentaux Grasshopper.xlsx (19 notions, 6 familles) – conservé sans modification"),
    ("Action de référence", "Compte rendu de session du 11/08/2026, §5.1 et §9.2 : « Comparer le tableau des "
                            "fondamentaux aux programmes RhinoForYou débutant et perfectionnement »."),
    ("", ""),
    ("SOURCES DÉPOUILLÉES", ""),
]
for code, label, fn in PROGS:
    lignes.append(("   " + code, label + "   —   " + fn))
lignes += [
    ("", ""),
    ("CONTENU DU CLASSEUR", ""),
    ("   Référentiel", "Feuille principale. %d lignes : %d notions issues des programmes de formation "
                       "+ 19 fondamentaux Magpie V1, classés par domaine puis par catégorie." % (len(ref), len(items))),
    ("   Fondamentaux V1", "Reprise à l'identique du tableau d'origine (19 notions), pour traçabilité."),
    ("   Contenu programmes", "Relevé exhaustif et verbatim des %d items des 6 programmes, dans l'ordre des "
                              "documents, rattachés à une ligne du référentiel." % len(rows)),
    ("   Synthèse", "Comptages par domaine, niveau, mode de validation et programme ; analyse des écarts."),
    ("   Listes", "Listes de valeurs utilisées par les menus déroulants du référentiel."),
    ("", ""),
    ("MÉTHODE", ""),
    ("   1", "Extraction automatique de tous les paragraphes de liste des 6 programmes (%d occurrences)." % len(rows)),
    ("   2", "Dédoublonnage par texte normalisé : %d items distincts." % len(items)),
    ("   3", "Classement de chaque item dans un domaine (10) et une catégorie, avec niveau et mode de validation proposés."),
    ("   4", "Intégration des 19 fondamentaux V1 et qualification de leur couverture par les programmes."),
    ("   5", "Construction de la matrice de couverture (une colonne par programme, valeur = jour · session)."),
    ("", ""),
    ("LÉGENDE — MODE DE VALIDATION", ""),
    ("   ExactOrderedList", "Ordre des éléments significatif."),
    ("   SetEquality", "Ordre indifférent, ensemble attendu."),
    ("   SingleValue", "Une valeur unique attendue."),
    ("   NumericTolerance", "Valeur numérique avec marge d'erreur."),
    ("   GeometryTolerance", "Comparaison géométrique avec tolérance."),
    ("   Conceptuel (QCM)", "Notion non vérifiable par comparaison de résultat : question ciblée."),
    ("", ""),
    ("LÉGENDE — ORIGINE", ""),
    ("   Programme de formation", "Item présent dans au moins un programme du catalogue."),
    ("   Fondamentaux V1 – couverte", "Notion V1 explicitement traitée par un item de programme."),
    ("   Fondamentaux V1 – partielle", "Notion V1 incluse dans un item de programme plus large."),
    ("   Fondamentaux V1 – hors programmes", "Notion V1 absente des programmes : à arbitrer (ajout au programme "
                                             "ou exercice Magpie seul)."),
    ("", ""),
    ("POINTS À ARBITRER", ""),
    ("   Écart principal", "Sur les 19 fondamentaux Magpie V1, 4 seulement sont explicitement couverts par les "
                           "programmes, 5 le sont partiellement et 10 en sont totalement absents (types et "
                           "conversion implicite, data matching, valeurs nulles, outils de texte, portes logiques)."),
    ("   Doublon de source", "« perfectionnement 6 jours » et « perfectionnement 6 jours (1) » ont un contenu "
                             "strictement identique ; l'indice A du 27/02/2025 y ajoute la rubrique "
                             "« Utilisation des plugins » (11 items)."),
    ("   Niveaux", "Les niveaux et modes de validation proposés sont une première affectation, à valider par "
                   "Jérémy CAROLUS avant production des exercices."),
    ("", ""),
    ("JOURNAL DES INDICES", ""),
]
for a, b in lignes:
    ws.append([a, b])
r0 = ws.max_row + 1
ws.cell(row=r0, column=1, value="Indice").font = HFONT
ws.cell(row=r0, column=2, value="Objet").font = HFONT
ws.cell(row=r0, column=3, value="Date").font = HFONT
ws.cell(row=r0, column=4, value="Auteur").font = HFONT
for j in range(1, 5):
    ws.cell(row=r0, column=j).fill = PatternFill("solid", fgColor=BLEU)
    ws.cell(row=r0, column=j).border = BORD
ws.cell(row=r0 + 1, column=1, value="Ind. A")
ws.cell(row=r0 + 1, column=2, value="Création. Fusion du tableau des fondamentaux V1 et de l'intégralité "
                                    "du contenu des 6 programmes de formation ; classement par domaine et "
                                    "catégorie ; matrice de couverture et analyse des écarts.")
ws.cell(row=r0 + 1, column=3, value=DATE_FR)
ws.cell(row=r0 + 1, column=4, value="C. THIERRY DE VILLE D'AVRAY")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
    for c in row:
        c.font = BFONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
for rr in range(2, ws.max_row + 1):
    v = ws.cell(row=rr, column=1).value
    if v and not v.startswith(" ") and ws.cell(row=rr, column=2).value in (None, ""):
        ws.cell(row=rr, column=1).font = Font(name="Calibri", size=10, bold=True, color=BLEU)
ws.sheet_view.showGridLines = False


# ============================================================ 2. REFERENTIEL
ws = wb.create_sheet("Référentiel")
HDR = ["N°", "ID", "Domaine", "Catégorie", "Notion", "Libellé source / description",
       "Niveau", "Origine / couverture",
       "D3\n3 j débutant", "P3\n3 j perf.", "P6\n6 j perf.", "P6b\n6 j perf. (1)",
       "P6A\n6 j perf. Ind A", "RG8\n8 j Rhino+GH", "Nb prog.",
       "Ordre V1", "Réf. croisée", "ValidationMode suggéré", "Type d'exercice",
       "Priorité de production", "Nb exercices Magpie prévus", "À réaliser par",
       "Statut", "Lien", "Notes"]
WID = [5, 9, 30, 30, 38, 62, 15, 32, 9, 9, 9, 9, 9, 9, 7, 8, 12, 20, 22, 11, 12, 18, 14, 26, 50]
entete(ws, "Référentiel des notions – Rhino / Grasshopper", len(HDR))
table_header(ws, 3, HDR, WID)

for r in ref:
    ws.append([
        r['ordre'], r['id'], r['dom'], r['cat'], r['notion'], r['desc'], r['niv'], r['origine'],
        r['cov'].get('D3', ''), r['cov'].get('P3', ''), r['cov'].get('P6', ''),
        r['cov'].get('P6b', ''), r['cov'].get('P6A', ''), r['cov'].get('RG8', ''),
        r['nbprog'] or '', r['ordre_v1'], r['reff'], r['mode'], r['typ'],
        PRIO[r['niv']], '', r['qui'], r['statut'], r['lien'], r['notes'],
    ])

last = ws.max_row
for i, rr in enumerate(range(4, last + 1)):
    fill = PatternFill("solid", fgColor=ACC[ref[i]['dom']])
    for j in range(1, len(HDR) + 1):
        c = ws.cell(row=rr, column=j)
        c.font = BFONT
        c.border = BORD
        c.alignment = Alignment(vertical="top", wrap_text=(j in (3, 4, 5, 6, 8, 17, 25)),
                                horizontal="center" if j in (1, 7, 9, 10, 11, 12, 13, 14, 15, 16, 20, 21, 23) else "left")
        if j in (3, 4):
            c.fill = fill
    if ref[i]['src'] == 'FND':
        ws.cell(row=rr, column=2).font = Font(name="Calibri", size=10, bold=True, color=BLEU)
    if ref[i]['origine'].endswith("(à ajouter)"):
        ws.cell(row=rr, column=8).font = Font(name="Calibri", size=10, bold=True, color="B00020")

ws.freeze_panes = "F4"
ws.auto_filter.ref = "A3:%s%d" % (get_column_letter(len(HDR)), last)
ws.sheet_view.showGridLines = False


# ============================================================ 3. FONDAMENTAUX V1
ws = wb.create_sheet("Fondamentaux V1")
entete(ws, "Tableau des fondamentaux Magpie V1 – reprise à l'identique (source J. CAROLUS)", 8)
H = ["Ordre de dépendance", "Catégorie", "Notion", "Description", "ValidationMode suggéré",
     "A réaliser par", "Statut", "Lien"]
table_header(ws, 3, H, [12, 24, 40, 80, 20, 18, 14, 26])
for f in fonds:
    ws.append([f['ordre'], f['cat'], f['notion'], f['desc'], f['mode'], f['qui'], f['statut'], f['lien']])
for rr in range(4, ws.max_row + 1):
    for j in range(1, 9):
        c = ws.cell(row=rr, column=j)
        c.font = BFONT
        c.border = BORD
        c.alignment = Alignment(vertical="top", wrap_text=(j in (2, 3, 4)),
                                horizontal="center" if j == 1 else "left")
ws.append([])
ws.append(["Légende ValidationMode :"])
for a, b in [("ExactOrderedList", "ordre des éléments compte"),
             ("SetEquality", "ordre indifférent, ensemble attendu"),
             ("SingleValue", "une valeur unique attendue"),
             ("NumericTolerance", "valeur numérique avec marge d'erreur"),
             ("GeometryTolerance", "comparaison géométrique avec tolérance")]:
    ws.append([a + " — " + b])
ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False


# ============================================================ 4. CONTENU PROGRAMMES
ws = wb.create_sheet("Contenu programmes")
entete(ws, "Relevé exhaustif du contenu des programmes de formation (verbatim)", 9)
H = ["N°", "Code", "Programme", "Jour", "Session", "Rubrique", "Item (verbatim)",
     "ID référentiel", "Domaine", "Catégorie"]
table_header(ws, 3, H, [5, 7, 34, 10, 14, 34, 78, 11, 30, 30])
idx_by_norm = {k: i for i, k in enumerate(items.keys(), 1)}
for n, r in enumerate(rows, 1):
    i = idx_by_norm[r['norm']]
    dom, cat = META[i][0], META[i][1]
    ws.append([n, r['prog'], r['prog_label'], r['jour'], r['session'], r['rubrique'],
               r['item'], id_by_idx[i], dom, cat])
for rr in range(4, ws.max_row + 1):
    for j in range(1, 11):
        c = ws.cell(row=rr, column=j)
        c.font = BFONT
        c.border = BORD
        c.alignment = Alignment(vertical="top", wrap_text=(j in (3, 6, 7, 9, 10)),
                                horizontal="center" if j in (1, 2, 8) else "left")
ws.freeze_panes = "D4"
ws.auto_filter.ref = "A3:J%d" % ws.max_row
ws.sheet_view.showGridLines = False


# ============================================================ 5. SYNTHESE
ws = wb.create_sheet("Synthèse")
entete(ws, "Synthèse et analyse des écarts", 6)
ws.column_dimensions['A'].width = 46
for col, w in zip("BCDEFG", [14, 14, 14, 14, 14, 14]):
    ws.column_dimensions[col].width = w
NREF = last  # derniere ligne du referentiel

def bloc(titre, headers, lignes, r):
    ws.cell(row=r, column=1, value=titre).font = Font(name="Calibri", size=11, bold=True, color=BLEU)
    r += 1
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=j, value=h)
        c.font = HFONT
        c.fill = PatternFill("solid", fgColor=BLEU)
        c.border = BORD
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1
    for l in lignes:
        for j, v in enumerate(l, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = BFONT
            c.border = BORD
            if j > 1:
                c.alignment = Alignment(horizontal="center")
        r += 1
    return r + 1


R = 3
lig = []
for d in DOMAINES:
    lig.append([d,
                '=COUNTIF(Référentiel!$C$4:$C$%d,$A%%d)' % NREF,
                '=COUNTIFS(Référentiel!$C$4:$C$%d,$A%%d,Référentiel!$H$4:$H$%d,"Programme de formation")' % (NREF, NREF),
                '=COUNTIFS(Référentiel!$C$4:$C$%d,$A%%d,Référentiel!$S$4:$S$%d,"P0")' % (NREF, NREF)])
start = R + 2
R = bloc("Répartition par domaine", ["Domaine", "Notions", "dont issues des programmes", "dont priorité P0"], lig, R)
for i in range(len(DOMAINES)):
    for j in (2, 3, 4):
        ws.cell(row=start + i, column=j).value = ws.cell(row=start + i, column=j).value % (start + i)

lig = [[n, '=COUNTIF(Référentiel!$G$4:$G$%d,$A%%d)' % NREF] for n in NIVEAUX]
start = R + 2
R = bloc("Répartition par niveau", ["Niveau", "Notions"], lig, R)
for i in range(len(NIVEAUX)):
    ws.cell(row=start + i, column=2).value = ws.cell(row=start + i, column=2).value % (start + i)

lig = [[m, '=COUNTIF(Référentiel!$R$4:$R$%d,$A%%d)' % NREF] for m in MODES]
start = R + 2
R = bloc("Répartition par mode de validation", ["ValidationMode", "Notions"], lig, R)
for i in range(len(MODES)):
    ws.cell(row=start + i, column=2).value = ws.cell(row=start + i, column=2).value % (start + i)

lig = [[t, '=COUNTIF(Référentiel!$S$4:$S$%d,$A%%d)' % NREF] for t in TYPES]
start = R + 2
R = bloc("Répartition par type d'exercice", ["Type d'exercice", "Notions"], lig, R)
for i in range(len(TYPES)):
    ws.cell(row=start + i, column=2).value = ws.cell(row=start + i, column=2).value % (start + i)

col_prog = dict(zip(PCODES, "IJKLMN"))
lig = []
for code, label, fn in PROGS:
    nb_occ = sum(1 for r in rows if r['prog'] == code)
    c = col_prog[code]
    lig.append([label, nb_occ,
                '=COUNTIF(Référentiel!$%s$4:$%s$%d,"<>")' % (c, c, NREF)])
R = bloc("Couverture par programme", ["Programme", "Items relevés (verbatim)", "Notions distinctes"], lig, R)

nc = [r for r in ref if r['src'] == 'FND' and r['origine'].endswith("(à ajouter)")]
pa = [r for r in ref if r['src'] == 'FND' and r['origine'].endswith("partielle")]
cv = [r for r in ref if r['src'] == 'FND' and r['origine'].endswith("les programmes")]
R = bloc("Écarts – fondamentaux Magpie V1 face aux programmes",
         ["Situation", "Nombre"],
         [["Notion V1 couverte par un item de programme", len(cv)],
          ["Notion V1 couverte partiellement", len(pa)],
          ["Notion V1 absente des programmes", len(nc)]], R)

ws.cell(row=R, column=1, value="Détail des notions V1 absentes des programmes de formation").font = \
    Font(name="Calibri", size=11, bold=True, color="B00020")
R += 1
for j, h in enumerate(["ID", "Catégorie", "Notion", "Niveau", "ValidationMode"], 1):
    c = ws.cell(row=R, column=j, value=h)
    c.font = HFONT
    c.fill = PatternFill("solid", fgColor="B00020")
    c.border = BORD
    c.alignment = Alignment(horizontal="center")
R += 1
for r in nc:
    for j, v in enumerate([r['id'], r['cat'], r['notion'], r['niv'], r['mode']], 1):
        c = ws.cell(row=R, column=j, value=v)
        c.font = BFONT
        c.border = BORD
    R += 1
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 46
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.sheet_view.showGridLines = False


# ============================================================ 6. LISTES
ws = wb.create_sheet("Listes")
entete(ws, "Listes de valeurs", 6)
cols = [("Domaines", DOMAINES), ("Niveaux", NIVEAUX), ("ValidationMode", MODES),
        ("Types d'exercice", TYPES), ("Priorités", ["P0", "P1", "P2"]),
        ("Statuts", ["Pas commencé", "En cours", "Terminé", "Bloqué"])]
for j, (t, vals) in enumerate(cols, 1):
    c = ws.cell(row=3, column=j, value=t)
    c.font = HFONT
    c.fill = PatternFill("solid", fgColor=BLEU)
    c.border = BORD
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width = 34
    for i, v in enumerate(vals, 4):
        cc = ws.cell(row=i, column=j, value=v)
        cc.font = BFONT
        cc.border = BORD
ws.sheet_view.showGridLines = False

# menus deroulants sur le referentiel
wsr = wb["Référentiel"]
for col, j, n in [("G", 2, len(NIVEAUX)), ("R", 3, len(MODES)), ("S", 4, len(TYPES)),
                  ("T", 5, 3), ("W", 6, 4)]:
    L = get_column_letter(j)
    dv = DataValidation(type="list", formula1="=Listes!$%s$4:$%s$%d" % (L, L, 3 + n), allow_blank=True)
    wsr.add_data_validation(dv)
    dv.add("%s4:%s%d" % (col, col, last))

wb.save(OUT)
print("OK ->", OUT)
print("Referentiel : %d lignes (%d items programme + %d fondamentaux V1)" % (len(ref), len(items), len(fonds)))
print("Contenu programmes : %d lignes verbatim" % len(rows))
print("Ecarts : %d couvertes / %d partielles / %d absentes" % (len(cv), len(pa), len(nc)))
