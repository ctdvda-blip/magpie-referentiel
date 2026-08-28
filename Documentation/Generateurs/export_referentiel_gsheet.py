# -*- coding: utf-8 -*-
"""Extrait l'onglet Referentiel pour alimentation d'un Google Sheet.

Retire les 10 colonnes de couverture par programme, demandees en moins :
    Origine / couverture, D3, P3, P6, P6b, P6A, RG8, Nb prog., Ordre V1, Ref. croisee

Applique la CHARTE GRAPHIQUE de la feuille d'origine "Fondamentaux Grasshopper" :
    - police Arial
    - ligne d'en-tete vert fonce, texte blanc gras
    - lignes de donnees en pastel, une teinte par domaine, reprenant les six
      couleurs d'origine completees par quatre harmonisees
    - blocs de synthese en pied de feuille, en-tetes bleu 2F5496 texte blanc
    - legende des modes de validation en italique

Produit trois formes du meme contenu, dans EXPORTS :
    Referentiel_pour_Google_Sheets.xlsx   -> a importer dans Google Sheets
    Referentiel_pour_Google_Sheets.csv    -> import CSV
    Referentiel_pour_Google_Sheets.tsv    -> a coller directement dans une feuille
"""
import os
import io
import csv
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ICI = os.path.dirname(os.path.abspath(__file__))
PROJET = os.path.abspath(os.path.join(ICI, "..", ".."))
SOURCE = os.path.join(PROJET, "Fondamentaux Grasshopper - IndA - 25-08-2026.xlsx")
DEST = os.path.join(PROJET, "EXPORTS")

A_RETIRER = [
    u"Origine / couverture",
    u"D3\n3 j débutant", u"P3\n3 j perf.", u"P6\n6 j perf.",
    u"P6b\n6 j perf. (1)", u"P6A\n6 j perf. Ind A", u"RG8\n8 j Rhino+GH",
    u"Nb prog.", u"Ordre V1", u"Réf. croisée",
]

# ---- charte reprise de la feuille d'origine -------------------------------
POLICE = "Arial"
VERT_ENTETE = "34694C"          # bandeau d'en-tete de la feuille source
BLEU_BLOC = "2F5496"            # en-tetes des blocs de synthese (exact, source)
GRIS_BORD = "BFC7D1"

# les six premieres teintes sont EXACTEMENT celles de la feuille d'origine
PASTELS = OrderedDict([
    (u"0 – Socle Rhino (prérequis)",                 "EFEFEF"),
    (u"1 – Environnement et principes Grasshopper",  "DCE6F1"),   # ex. Listes
    (u"2 – Données et logique",                      "EAD1DC"),   # ex. Comportements implicites
    (u"3 – Géométrie paramétrique",                  "FDE9D9"),   # ex. Types
    (u"4 – Mesures, quantitatifs et export",         "FFF2CC"),   # ex. Outils de texte
    (u"5 – Méthode, performance et évènements",      "D9D2E9"),   # ex. Portes logiques
    (u"6 – Algorithmique avancée",                   "D9EAD3"),   # ex. Arbres
    (u"7 – Développement, scripting et API",         "D0E0E3"),
    (u"8 – Interfaces, web et interopérabilité",     "F4CCCC"),
    (u"9 – Aide à la fabrication",                   "E6E6C8"),
])

LEGENDE = [
    (u"ExactOrderedList", u"ordre des éléments compte"),
    (u"SetEquality", u"ordre indifférent, ensemble attendu"),
    (u"SingleValue", u"une valeur unique attendue"),
    (u"NumericTolerance", u"valeur numérique avec marge d'erreur"),
    (u"GeometryTolerance", u"comparaison géométrique avec tolérance"),
    (u"Conceptuel (QCM)", u"notion non vérifiable par comparaison de résultat"),
]

LARGEURS = [5, 10, 34, 32, 40, 70, 15, 22, 24, 12, 13, 18, 14, 26, 55]

THIN = Side(style="thin", color=GRIS_BORD)
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def normaliser(v):
    return (v or u"").replace(u"\n", u" ").strip()


def lire():
    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    ws = wb[u"Référentiel"]
    entetes = [c.value for c in ws[3]]
    retirer = set(i for i, e in enumerate(entetes) if e in A_RETIRER)
    manquantes = [a for a in A_RETIRER if a not in entetes]
    if manquantes:
        print(u"!! colonnes demandées introuvables : %s"
              % u", ".join(m.replace(u"\n", u" ") for m in manquantes))
    gardees = [i for i in range(len(entetes)) if i not in retirer]

    lignes = [[normaliser(entetes[i]) for i in gardees]]
    for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if r[0] is None:
            continue
        ligne = []
        for i in gardees:
            v = r[i] if i < len(r) else None
            if v is None:
                v = u""
            elif isinstance(v, float) and v == int(v):
                v = int(v)
            ligne.append(v)
        lignes.append(ligne)
    return lignes, len(retirer)


def texte(ws, r, c, v, taille=10, gras=False, italique=False, blanc=False,
          fond=None, centre=False, retour=False, bord=True):
    cel = ws.cell(row=r, column=c, value=v)
    cel.font = Font(name=POLICE, size=taille, bold=gras, italic=italique,
                    color="FFFFFF" if blanc else "000000")
    if fond:
        cel.fill = PatternFill("solid", fgColor=fond)
    cel.alignment = Alignment(horizontal="center" if centre else "left",
                              vertical="center" if centre else "top",
                              wrap_text=retour)
    if bord:
        cel.border = BORD
    return cel


def ecrire_xlsx(lignes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = u"Référentiel"
    entetes = lignes[0]
    n = len(entetes)

    # --- en-tete
    for j, v in enumerate(entetes, 1):
        texte(ws, 1, j, v, taille=10, gras=True, blanc=True,
              fond=VERT_ENTETE, centre=True, retour=True)
    ws.row_dimensions[1].height = 30

    # --- donnees, pastel par domaine
    idom = entetes.index(u"Domaine")
    for i, l in enumerate(lignes[1:], 2):
        fond = PASTELS.get(l[idom], None)
        for j, v in enumerate(l, 1):
            texte(ws, i, j, v, taille=10, fond=fond,
                  centre=j in (1, 7, 10, 11),
                  retour=j in (3, 4, 5, 6, n))

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(n), len(lignes))
    for j, w in enumerate(LARGEURS[:n], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # --- legende, comme dans la feuille d'origine
    r = len(lignes) + 2
    texte(ws, r, 1, u"Légende ValidationMode :", gras=True, bord=False)
    for k, (nom, desc) in enumerate(LEGENDE, 1):
        texte(ws, r + k, 1, u"%s — %s" % (nom, desc), italique=True, bord=False)

    # --- synthese par domaine
    r = r + len(LEGENDE) + 3
    texte(ws, r, 1, u"Domaine", gras=True, blanc=True, fond=BLEU_BLOC, centre=True)
    texte(ws, r, 2, u"Nb notions", gras=True, blanc=True, fond=BLEU_BLOC, centre=True)
    for k, dom in enumerate(PASTELS, 1):
        texte(ws, r + k, 1, dom, fond=PASTELS[dom])
        # nom INTERNE de la fonction : le format de fichier n'accepte pas le
        # libelle francais NB.SI, qui donnerait #NOM? a l'ouverture.
        texte(ws, r + k, 2,
              u'=COUNTIF($C$2:$C$%d,$A$%d)' % (len(lignes), r + k), centre=True)

    # --- synthese par mode de validation
    r = r + len(PASTELS) + 3
    imode = entetes.index(u"ValidationMode suggéré") + 1
    col = get_column_letter(imode)
    texte(ws, r, 1, u"Mode de validation", gras=True, blanc=True,
          fond=BLEU_BLOC, centre=True)
    texte(ws, r, 2, u"Nb", gras=True, blanc=True, fond=BLEU_BLOC, centre=True)
    for k, (nom, _) in enumerate(LEGENDE, 1):
        texte(ws, r + k, 1, nom)
        texte(ws, r + k, 2,
              u'=COUNTIF($%s$2:$%s$%d,$A$%d)' % (col, col, len(lignes), r + k),
              centre=True)

    chemin = os.path.join(DEST, u"Referentiel_pour_Google_Sheets.xlsx")
    wb.save(chemin)
    return chemin


def main():
    lignes, nb_retirees = lire()
    if not os.path.isdir(DEST):
        os.makedirs(DEST)
    print(u"Colonnes retirées : %d" % nb_retirees)
    print(u"Colonnes conservées : %d" % len(lignes[0]))
    print(u"Lignes de données : %d" % (len(lignes) - 1))

    for ext, sep in ((u"csv", u","), (u"tsv", u"\t")):
        chemin = os.path.join(DEST, u"Referentiel_pour_Google_Sheets." + ext)
        fh = io.open(chemin, "w", encoding="utf-8-sig", newline="")
        w = csv.writer(fh, delimiter=str(sep), quoting=csv.QUOTE_MINIMAL,
                       lineterminator="\n")
        for l in lignes:
            w.writerow([u"%s" % x for x in l])
        fh.close()
        print(u"écrit : %s" % chemin)

    print(u"écrit : %s" % ecrire_xlsx(lignes))
    print(u"")
    print(u"Charte appliquée : Arial, en-tête vert %s, pastels par domaine "
          u"(6 teintes reprises de la feuille d'origine), blocs de synthèse "
          u"bleu %s." % (VERT_ENTETE, BLEU_BLOC))


if __name__ == "__main__":
    main()
