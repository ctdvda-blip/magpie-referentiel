# -*- coding: utf-8 -*-
"""Genere l'application HTML de consultation MAGPIE.

Un fichier unique, autonome, place a la racine du projet. Il presente :
    - le referentiel unifie (116 notions), filtrable
    - les 49 exercices du lot A, avec le contenu integral de leur fiche
    - les liens vers les definitions Grasshopper et les telechargements
      PDF / Word de chaque fiche

Le logo RhinoForYou est incorpore en base64 : l'application reste lisible
meme deplacee, tant qu'elle garde le dossier EXERCICES a cote d'elle.
"""
import os
import sys
import io
import json
import base64

ICI = os.path.dirname(os.path.abspath(__file__))
for p in (ICI, os.path.join(ICI, "GH")):
    if p not in sys.path:
        sys.path.insert(0, p)

import openpyxl
from lots import TOUS as LOT_A, dossier_de
from skill_a import SKILL
# enonce_origine est deja pose par le registre pour le lot A ; les autres lots
# n'ont pas d'enonce anterieur a conserver.

PROJET = os.path.abspath(os.path.join(ICI, "..", ".."))
CLASSEUR = os.path.join(PROJET, "Fondamentaux Grasshopper - IndB - 26-08-2026.xlsx")
DOSSIER_EX = "EXERCICES/LOT A - Composants natifs"
DOSSIER_IA = "EXERCICES/LOT IA - IA et assistance generative"
SORTIE = os.path.join(PROJET, "MAGPIE - Application.html")
# GitHub Pages sert « index.html » a la racine : on ecrit les deux noms
# depuis la meme source, pour qu'ils ne puissent pas diverger.
SORTIE_INDEX = os.path.join(PROJET, "index.html")

# --plat <chemin> : dossiers d'exercice nommes par leur seul identifiant,
# et sortie ecrite a l'endroit indique. Sert a la publication.
PLAT = "--plat" in sys.argv
# --livrables <racine> : sonder la presence des fichiers dans CETTE arborescence
# plutot que dans le projet local. Sans cela, une publication qui n'embarque pas
# encore les PDF proposerait quand meme le bouton — et le lien serait mort.
RACINE_LIVRABLES = None
if "--livrables" in sys.argv:
    _k = sys.argv.index("--livrables")
    if _k + 1 < len(sys.argv):
        RACINE_LIVRABLES = os.path.abspath(sys.argv[_k + 1])
# --protege LOGIN:MOTDEPASSE  ajoute un ecran d'entree devant l'application.
# Le mot de passe n'est JAMAIS ecrit dans la page : on n'y met qu'un sel
# aleatoire et l'empreinte PBKDF2 du mot de passe. La page ne peut donc pas
# reveler le secret — elle peut seulement verifier celui qu'on lui presente.
PROTEGE = None
if "--protege" in sys.argv:
    import hashlib, binascii, os as _os
    _j = sys.argv.index("--protege")
    if _j + 1 < len(sys.argv):
        _login, _mdp = sys.argv[_j + 1].split(":", 1)
        _sel = _os.urandom(16)
        _emp = hashlib.pbkdf2_hmac("sha256", _mdp.encode("utf-8"), _sel,
                                   200000, 32)
        PROTEGE = (_login,
                   binascii.hexlify(_sel).decode(),
                   binascii.hexlify(_emp).decode())
        del _mdp
if PLAT:
    _i = sys.argv.index("--plat")
    if _i + 1 < len(sys.argv):
        SORTIE_INDEX = os.path.abspath(sys.argv[_i + 1])
        SORTIE = SORTIE_INDEX

LOGOS = [
    r"C:\Users\charl\Documents\PROJETS\IA\IMAGES\LOGO-RHINOFORYOU-XXL-300x300.png",
    r"C:\Users\charl\Downloads\LOGO-RHINOFORYOU-XXL-300x300.png",
    r"C:\Users\charl\OneDrive\Bureau\T CONCEPT 3D\IMAGES\LOGOS\LogoRhinoforyou.png",
]

VERSION = "v0.3-260826"
INDICE = "Ind. B"
DATE = "26/08/2026"

# Une couleur par domaine. Les clefs suivent la numerotation de 1 a 11 : elles
# etaient restees sur l'ancienne numerotation commencant a 0, et toutes les
# tuiles retombaient donc sur la couleur de repli.
COULEURS = {
    u"1 – Socle Rhino (prérequis)": "#7c8894",
    u"2 – Environnement et principes Grasshopper": "#5b8cc4",
    u"3 – Données et logique": "#5aa469",
    u"4 – Géométrie paramétrique": "#d99a4e",
    u"5 – Mesures, quantitatifs et export": "#c9b03a",
    u"6 – Méthode, performance et évènements": "#b56b96",
    u"7 – Algorithmique avancée": "#8a7cc2",
    u"8 – Développement, scripting et API": "#4fa3a8",
    u"9 – Interfaces, web et interopérabilité": "#cc6666",
    u"10 – Aide à la fabrication": "#9a9a5e",
    u"11 – IA et assistance générative": "#7f5fc9",
}


PORTAIL = u"""

/* ------------------------------------------------------------------ portail
   Ecran d'entree. ATTENTION, ET C'EST ECRIT AUSSI DANS LE README :
   le mot de passe n'est PAS dans cette page — on n'y trouve que son empreinte
   PBKDF2 et un sel, d'ou l'on ne peut pas le remonter. En revanche, ce portail
   ne protege pas les FICHIERS : le depot est public, et chaque definition,
   fiche ou corrige reste accessible par son adresse directe, portail ou pas.
   Il garde le hall, pas les portes.
------------------------------------------------------------------------- */
(function(){
  const LOGIN = %s, SEL = %s, EMPREINTE = %s, TOURS = 200000;

  const octets = h => Uint8Array.from(h.match(/../g).map(x => parseInt(x, 16)));
  const hexa = b => [...new Uint8Array(b)]
    .map(x => x.toString(16).padStart(2, '0')).join('');

  async function verifier(mdp){
    const k = await crypto.subtle.importKey(
      'raw', new TextEncoder().encode(mdp), 'PBKDF2', false, ['deriveBits']);
    const bits = await crypto.subtle.deriveBits(
      {name:'PBKDF2', salt: octets(SEL), iterations: TOURS, hash:'SHA-256'},
      k, 256);
    return hexa(bits) === EMPREINTE;
  }
  const CLE = 'magpie-entree';
  const ok = () => { try { return sessionStorage.getItem(CLE) === '1'; }
                     catch(e){ return false; } };
  if (ok()) return;

  const v = document.createElement('div');
  v.id = 'portail';
  v.innerHTML = `<form id="pf">
      <h1>MAGPIE</h1>
      <p>Référentiel et exercices Grasshopper — accès réservé.</p>
      <label>Identifiant<input id="pl" autocomplete="username" autofocus></label>
      <label>Mot de passe<input id="pm" type="password" autocomplete="current-password"></label>
      <button type="submit">Entrer</button>
      <div id="pe"></div>
    </form>`;
  document.body.appendChild(v);
  document.body.style.overflow = 'hidden';

  document.getElementById('pf').onsubmit = async ev => {
    ev.preventDefault();
    const l = document.getElementById('pl').value.trim();
    const m = document.getElementById('pm').value;
    document.getElementById('pe').textContent = 'Vérification…';
    if (l === LOGIN && await verifier(m)){
      try { sessionStorage.setItem(CLE, '1'); } catch(e){}
      v.remove();
      document.body.style.overflow = '';
    } else {
      document.getElementById('pe').textContent =
        'Identifiant ou mot de passe incorrect.';
      document.getElementById('pm').value = '';
      document.getElementById('pm').focus();
    }
  };
})();
"""

PORTAIL_CSS = u"""
#portail{position:fixed;inset:0;z-index:9999;background:var(--fond);
 display:flex;align-items:center;justify-content:center;padding:24px}
#portail form{background:var(--carte);border:1px solid var(--trait);
 border-radius:14px;box-shadow:var(--ombre);padding:34px 32px;width:min(380px,100%);
 display:flex;flex-direction:column;gap:14px}
#portail h1{margin:0;font-size:30px;letter-spacing:.16em;color:var(--bleu);text-align:center}
#portail p{margin:0 0 6px;color:var(--doux);font-size:13.5px;text-align:center}
#portail label{display:flex;flex-direction:column;gap:5px;font-size:13px;
 font-weight:600;color:var(--doux)}
#portail input{padding:9px 11px;border:1px solid var(--trait);border-radius:8px;
 font:15px "Segoe UI",Calibri,sans-serif;color:var(--texte);background:#fff}
#portail input:focus{outline:2px solid var(--bleu);outline-offset:1px;border-color:var(--bleu)}
#portail button{margin-top:6px;padding:10px;border:0;border-radius:8px;
 background:var(--bleu);color:#fff;font:600 15px "Segoe UI",Calibri,sans-serif;cursor:pointer}
#portail button:hover{background:#24486b}
#portail #pe{min-height:18px;color:var(--rouge);font-size:13px;text-align:center}
"""

MANUELS = {}
try:
    import recipes_a1, recipes_a2, recipes_a3, recipes_a4
    for m in (recipes_a1, recipes_a2, recipes_a3, recipes_a4):
        for k, v in m.R.items():
            if v.get("manuel"):
                MANUELS[k] = v["manuel"]
except Exception as e:
    print(u"  (réglages manuels non repris : %s)" % e)


def logo_b64():
    for p in LOGOS:
        if os.path.isfile(p):
            return base64.b64encode(io.open(p, "rb").read()).decode("ascii")
    print(u"  !! logo introuvable, l'en-tête affichera le texte seul")
    return None


def lire_referentiel():
    wb = openpyxl.load_workbook(CLASSEUR, data_only=False)
    ws = wb[u"Référentiel"]
    entetes = [c.value for c in ws[3]]
    out = []
    for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if r[0] is None:
            continue
        d = {}
        for i, e in enumerate(entetes):
            v = r[i] if i < len(r) else None
            d[e] = u"" if v is None else v
        out.append(d)
    return out


def correspondance():
    """Ancien identifiant (FND-/PRG-) -> nouvel identifiant unifie REF-."""
    import csv as _csv
    c = os.path.join(PROJET, "EXPORTS", "Correspondance_identifiants_IndA_vers_IndB.csv")
    m = {}
    if os.path.isfile(c):
        fh = io.open(c, encoding="utf-8-sig")
        for i, l in enumerate(_csv.reader(fh, delimiter=";")):
            if i and len(l) >= 2:
                m[l[0].strip()] = l[1].strip()
        fh.close()
    return m


def exercices(corr, ref):
    par_id = dict((r[u"ID"], r) for r in ref)
    out = []
    for e in LOT_A:
        # En mode publication, les dossiers portent le seul identifiant : sur
        # Windows, « A-02 Construire un point par coordonnées/Illustrations/... »
        # depasse la limite de 260 caracteres des qu'on clone dans un chemin un
        # peu profond, et le clone echoue.
        dossier = e["id"] if PLAT else u"%s %s" % (e["id"], e["titre"])
        # Chaque lot a son dossier : le chemin ne peut pas etre suppose.
        racine = dossier_de(e["id"])
        # On ne propose au telechargement que ce qui existe : un lot dont les
        # livrables restent a produire ne doit pas afficher de liens morts.
        if RACINE_LIVRABLES:
            base = os.path.join(RACINE_LIVRABLES, racine, dossier)
        else:
            base = os.path.join(PROJET, racine,
                                u"%s %s" % (e["id"], e["titre"]))
            if not os.path.isdir(base):
                base = os.path.join(PROJET, racine, e["id"])
        dispo = {}
        for cle, nom in ((u"pdf", u"%s_fiche.pdf"),
                         (u"docx", u"%s_fiche.docx"),
                         (u"docx_sujet", u"%s_fiche_sujet.docx"),
                         (u"gh_sujet", u"%s_sujet.gh"),
                         (u"gh_complet", u"%s_complet.gh")):
            dispo[cle] = os.path.isfile(os.path.join(base, nom % e["id"]))
        dispo[u"images"] = os.path.isfile(os.path.join(
            base, u"Illustrations", u"web", u"%s_canvas_sujet.jpg" % e["id"]))
        anciens = [x.strip() for x in e["ref"].split(",") if x.strip()]
        nouveaux = [corr.get(a, a) for a in anciens]
        doms, cats = [], []
        for n in nouveaux:
            r = par_id.get(n)
            if not r:
                continue
            if r[u"Domaine"] not in doms:
                doms.append(r[u"Domaine"])
            if r[u"Catégorie"] not in cats:
                cats.append(r[u"Catégorie"])
        dom = doms[0] if doms else u""
        out.append({
            "id": e["id"], "titre": e["titre"], "them": e["them"],
            "ref": u", ".join(nouveaux), "dom": dom,
            "doms": doms, "cats": cats,
            "niv": e["niv"], "duree": e["duree"],
            "prereq": e["prereq"], "mode": e["mode"], "tol": e["tol"],
            "nb": e["nb"], "gamif": e.get("gamif", u"—"),
            "obj": e["obj"], "comp": e["comp"], "enonce": e["enonce"],
            "depart": e["depart"], "att": e["att"], "bareme": e["bareme"],
            "etapes": e["etapes"], "pieges": e["pieges"], "var": e["var"],
            "manuel": MANUELS.get(e["id"], []),
            "dossier": dossier,
            "racine": racine,
            "dispo": dispo,
            # --- champs issus de la skill de conception ------------------
            "competence": e.get("competence") or u"",
            "bloom": e.get("bloom") or u"",
            "contexte": e.get("contexte") or u"",
            "erreur": e.get("erreur") or u"",
            "verdict": e.get("verdict") or u"competence",
            "charniere": e.get("charniere") or u"",
            "enonce_origine": e.get("enonce_origine") or u"",
            "donnees_note": e.get("donnees_note") or u"",
            "limite": e.get("limite") or u"",
            "alerte": e.get("alerte") or u"",
        })
    return out


CSS = u"""
*{box-sizing:border-box}
:root{
 --fond:#f4f6f8; --carte:#ffffff; --texte:#1f2a37; --doux:#5b6775;
 --trait:#dfe4ea; --bleu:#2e5c8a; --bleu-clair:#eaf1f8; --ambre:#fdf3d6;
 --h-entete:68px; --rouge:#b00020; --ombre:0 1px 3px rgba(16,24,40,.08),0 1px 2px rgba(16,24,40,.04);
}
html,body{margin:0;padding:0}
body{background:var(--fond);color:var(--texte);
 font:15px/1.55 "Segoe UI",Calibri,system-ui,sans-serif}
a{color:var(--bleu)}
header.top{position:sticky;top:0;z-index:20;background:#fff;border-bottom:1px solid var(--trait);
 display:flex;align-items:center;gap:16px;padding:10px 22px;box-shadow:var(--ombre)}
header.top img{width:46px;height:46px;object-fit:contain;border-radius:6px}
.accueil-logo{display:flex;align-items:center;gap:16px;background:none;border:0;
 padding:0;margin:0;font:inherit;color:inherit;cursor:pointer;text-align:left;
 border-radius:8px}
.accueil-logo:hover{opacity:.75}
.accueil-logo:focus-visible{outline:2px solid var(--bleu);outline-offset:3px}
.marque{font-weight:700;font-size:15px;letter-spacing:.02em}
.marque small{display:block;font-weight:400;font-size:12px;color:var(--doux)}
nav{margin-left:26px;display:flex;gap:4px}
nav button{border:0;background:transparent;font:inherit;color:var(--doux);
 padding:8px 15px;border-radius:8px;cursor:pointer}
nav button:hover{background:var(--fond)}
nav button.actif{background:var(--bleu);color:#fff}
.version{margin-left:auto;font-size:11.5px;color:var(--doux);text-align:right;line-height:1.35}
main{max-width:1360px;margin:0 auto;padding:26px 22px 70px}
.vue{display:none}.vue.actif{display:block}

.hero{text-align:center;padding:34px 20px 24px}
.hero h1{margin:0;font-size:62px;letter-spacing:.16em;color:var(--bleu);font-weight:800}
.hero .sous{margin:8px 0 0;color:var(--doux);font-size:17px}
.chiffres{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px;margin:30px 0}
.chiffre{background:var(--carte);border:1px solid var(--trait);border-radius:12px;
 padding:18px;text-align:center;box-shadow:var(--ombre)}
.chiffre b{display:block;font-size:33px;color:var(--bleu);line-height:1.1}
.chiffre span{font-size:13px;color:var(--doux)}
.blocs{display:grid;grid-template-columns:repeat(auto-fit,minmax(290px,1fr));gap:16px}
.bloc{background:var(--carte);border:1px solid var(--trait);border-radius:12px;padding:20px;
 box-shadow:var(--ombre)}
.bloc h3{margin:0 0 8px;font-size:16px}
.bloc p{margin:0 0 12px;color:var(--doux);font-size:14px}

.theme h2{margin:34px 0 4px;font-size:21px}
.theme .aide{margin:0 0 14px;color:var(--doux);font-size:14px}
.fil{display:flex;flex-wrap:wrap;align-items:center;gap:8px;margin:0 0 16px;font-size:14px}
.fil button{border:0;background:var(--bleu-clair);color:var(--bleu);font:inherit;
 padding:6px 13px;border-radius:999px;cursor:pointer}
.fil button:hover{filter:brightness(.96)}
.fil span.sep{color:var(--doux)}
.fil span.ici{padding:6px 13px;border-radius:999px;background:var(--bleu);color:#fff}
.tuiles{display:grid;grid-template-columns:repeat(auto-fill,minmax(268px,1fr));gap:13px}
.tuile{background:var(--carte);border:1px solid var(--trait);border-radius:12px;
 padding:15px 16px;cursor:pointer;box-shadow:var(--ombre);display:flex;gap:12px;
 align-items:flex-start;transition:transform .12s}
.tuile:hover{transform:translateY(-2px)}
.tuile .bande{width:6px;align-self:stretch;border-radius:4px;flex:none}
.tuile b{display:block;font-size:14.5px;margin-bottom:5px;line-height:1.35}
.tuile .cpt{font-size:12.5px;color:var(--doux)}
.tuile .fleche{margin-left:auto;color:var(--doux);font-size:18px;align-self:center}
.duo{display:grid;grid-template-columns:1fr 1fr;gap:20px;align-items:start}
@media(max-width:980px){.duo{grid-template-columns:1fr}}
.duo h3{margin:0 0 10px;font-size:15px;color:var(--bleu)}
.duo .cartes{grid-template-columns:1fr}
.duo .carte h4{margin:4px 0 0}
.mini{background:var(--carte);border:1px solid var(--trait);border-radius:12px;
 padding:6px 0;box-shadow:var(--ombre)}
.mini .l{padding:9px 15px;border-top:1px solid var(--trait);font-size:13.5px}
.mini .l:first-child{border-top:0}
.mini .l b{display:block;font-size:14px}
.mini .l span{color:var(--doux);font-size:12.5px}

.barre{display:flex;flex-wrap:wrap;gap:10px;align-items:center;margin-bottom:16px}
input[type=search],select{font:inherit;padding:9px 12px;border:1px solid var(--trait);
 border-radius:9px;background:#fff;color:var(--texte)}
input[type=search]{flex:1;min-width:230px}
.compte{color:var(--doux);font-size:13px}

/* Le cadre est le conteneur defilant : c'est LUI qui porte overflow, et la
   ligne d'en-tetes se colle a son bord haut. Un overflow sur la table elle-meme
   annulerait le position:sticky de ses cellules d'en-tete — c'est ce qui les
   faisait flotter au milieu du tableau. */
.cadre-table{border:1px solid var(--trait);border-radius:12px;overflow:auto;
 box-shadow:var(--ombre);background:var(--carte);
 max-height:calc(100vh - var(--h-entete) - 150px)}
/* la valeur ci-dessus n'est qu'un repli : mesurerEntete() la remplace */
table{width:100%;border-collapse:separate;border-spacing:0;background:var(--carte)}
th{position:sticky;top:0;background:var(--bleu);color:#fff;text-align:left;
 padding:10px;font-size:12.5px;font-weight:600;z-index:5;
 box-shadow:inset 0 -1px 0 rgba(255,255,255,.18)}
td{padding:9px 10px;border-top:1px solid var(--trait);font-size:13.5px;vertical-align:top}
tr:hover td{background:#f8fafc}
.pastille{display:inline-block;width:9px;height:9px;border-radius:50%;margin-right:7px}
.mono{font-family:Consolas,monospace;font-size:12.5px;color:var(--doux)}
.etiq{display:inline-block;padding:2px 9px;border-radius:999px;font-size:11.5px;
 background:var(--bleu-clair);color:var(--bleu);white-space:nowrap}

.cartes{display:grid;grid-template-columns:repeat(auto-fill,minmax(310px,1fr));gap:14px}
.carte{background:var(--carte);border:1px solid var(--trait);border-radius:12px;padding:16px;
 cursor:pointer;box-shadow:var(--ombre);border-left:5px solid var(--bleu);transition:transform .12s}
.carte:hover{transform:translateY(-2px)}
.carte .cid{font-family:Consolas,monospace;font-size:12px;color:var(--doux)}
.carte h4{margin:4px 0 8px;font-size:15.5px}
.carte .meta{display:flex;flex-wrap:wrap;gap:6px;margin-top:10px}

.fiche{background:var(--carte);border:1px solid var(--trait);border-radius:14px;
 padding:26px;box-shadow:var(--ombre)}
.fiche h2{margin:0 0 4px;font-size:26px;color:var(--bleu)}
.fiche .them{color:var(--doux);margin-bottom:18px;font-size:14px}
.infos{width:100%;border:1px solid var(--trait);border-radius:10px;overflow:hidden;margin-bottom:20px}
.infos td{border-top:1px solid var(--trait);padding:7px 11px;font-size:13px}
.infos tr:first-child td{border-top:0}
.infos td:first-child{background:#f1f5f9;font-weight:600;width:230px}
h3.sect{margin:26px 0 12px;padding-bottom:7px;border-bottom:2px solid var(--bleu);
 color:var(--bleu);font-size:17px}
h4.sous{margin:18px 0 6px;font-size:14px;color:var(--doux);text-transform:uppercase;
 letter-spacing:.04em}
.encadre{background:var(--bleu-clair);border:1px solid #cfe0f0;border-radius:10px;
 padding:14px 16px;margin:6px 0 14px;font-size:15px}
.encadre.ambre{background:var(--ambre);border-color:#efe0b0}
.encadre.rouge{background:#fdecef;border-color:#f3ccd4}
.note{color:var(--doux);font-size:13.5px;font-style:italic;margin:6px 0 10px}
.cell-exos{white-space:nowrap}
.bt-exo{display:inline-block;margin:1px 3px 1px 0;padding:3px 8px;font:600 11.5px/1.3
 "Segoe UI",Calibri,sans-serif;color:#fff;background:var(--bleu);border:0;
 border-radius:5px;cursor:pointer;letter-spacing:.02em}
.bt-exo:hover{background:#24486b}
.bt-exo:focus-visible{outline:2px solid var(--bleu);outline-offset:2px}
.bt-exo.charniere{background:#8a6d1f}
.bt-exo.charniere:hover{background:#6f570f}
.vide-cell{color:var(--doux)}
.note b{font-style:normal}
figure{margin:14px 0}
figure img{width:100%;border:1px solid var(--trait);border-radius:10px;background:#fff}
figcaption{font-size:12.5px;color:var(--doux);text-align:center;margin-top:6px;font-style:italic}
ol.etapes{padding-left:0;list-style:none;counter-reset:e}
ol.etapes li{counter-increment:e;margin:0 0 9px;padding-left:78px;position:relative}
ol.etapes li::before{content:"Étape " counter(e);position:absolute;left:0;top:0;
 font-weight:700;color:var(--bleu);font-size:13px}
ul.puces{padding-left:20px;margin:6px 0}
ul.puces li{margin:4px 0}
.telech{display:flex;flex-wrap:wrap;gap:10px;margin:22px 0 4px;padding-top:18px;
 border-top:1px solid var(--trait)}
.bt{display:inline-flex;align-items:center;gap:7px;padding:9px 15px;border-radius:9px;
 border:1px solid var(--trait);background:#fff;color:var(--texte);text-decoration:none;
 font-size:13.5px;cursor:pointer;font-family:inherit}
.bt:hover{background:var(--fond)}
.bt.plein{background:var(--bleu);color:#fff;border-color:var(--bleu)}
.bt.plein:hover{opacity:.9}
.interrupteur{display:flex;align-items:center;gap:10px;background:var(--ambre);
 border:1px solid #efe0b0;border-radius:10px;padding:11px 15px;margin:8px 0 4px;font-size:14px}
.interrupteur input{width:17px;height:17px;cursor:pointer}
#corrige.masque{display:none}
.retour{background:none;border:0;color:var(--bleu);cursor:pointer;font:inherit;
 padding:0;margin-bottom:14px}
.vide{text-align:center;color:var(--doux);padding:50px}
header.top{flex-wrap:wrap;row-gap:8px}
@media(max-width:860px){
  header.top{padding:8px 14px;gap:10px}
  header.top img{width:34px;height:34px}
  .marque small{display:none}
  .marque{font-size:15px}
  header.top nav{order:3;width:100%}
  header.top .version{margin-left:auto}
}
@media print{
 header.top,nav,.barre,.telech,.retour,.interrupteur{display:none!important}
 .cadre-table{max-height:none;overflow:visible;box-shadow:none}
 th{position:static}
 body{background:#fff}main{max-width:none;padding:0}
 .fiche{border:0;box-shadow:none;padding:0}
 #corrige.masque{display:block!important}
}
"""


def main():
    ref = lire_referentiel()
    corr = correspondance()
    exos = exercices(corr, ref)
    orphelins = [e["id"] for e in exos if not e["dom"]]
    if orphelins:
        print(u"  !! exercices sans domaine résolu : %s" % u", ".join(orphelins))
    logo = logo_b64()
    # Tri naturel : un tri alphabetique placerait « 10 » et « 11 » entre
    # « 1 » et « 2 ». On trie sur le numero de tete, pas sur la chaine.
    def _rang(d):
        n = u""
        for c in d:
            if c.isdigit():
                n += c
            else:
                break
        return (int(n) if n else 999, d)

    domaines = sorted(set(r[u"Domaine"] for r in ref), key=_rang)
    themes = sorted(set(e["them"] for e in exos))

    donnees = {
        "ref": ref, "exos": exos, "couleurs": COULEURS,
        "domaines": domaines, "themes": themes, "dossier": DOSSIER_EX,
        "version": VERSION, "indice": INDICE, "date": DATE,
    }
    js = json.dumps(donnees, ensure_ascii=False).replace(u"</", u"<\\/")

    logo_html = (u'<img src="data:image/png;base64,%s" alt="RhinoForYou">' % logo) if logo else u""
    favicon = (u'<link rel="icon" href="data:image/png;base64,%s">' % logo) if logo else u""

    html = u"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>MAGPIE — Référentiel et exercices Grasshopper</title>
%s
<style>%s</style></head><body>

<header class="top">
  <button class="accueil-logo" id="retour-accueil" type="button"
          title="Revenir à l'accueil et aux catégories">
    %s
    <span class="marque">RhinoForYou<small>Magpie — exercices Grasshopper</small></span>
  </button>
  <nav>
    <button data-vue="accueil" class="actif">Accueil</button>
    <button data-vue="referentiel">Référentiel</button>
    <button data-vue="exercices">Exercices</button>
  </nav>
  <div class="version">%s — %s<br>%s</div>
</header>

<main>
  <section id="vue-accueil" class="vue actif">
    <div class="hero">
      <h1>MAGPIE</h1>
      <p class="sous">Référentiel des notions Rhino / Grasshopper et bibliothèque d'exercices autocorrigés</p>
    </div>
    <div class="chiffres" id="chiffres"></div>
    <section class="theme">
      <h2>Accès thématique</h2>
      <p class="aide">Choisissez un domaine, puis une catégorie, pour voir les notions
         concernées et les exercices qui s'y rattachent.</p>
      <div class="fil" id="fil"></div>
      <div id="theme"></div>
    </section>

    <div class="blocs" style="margin-top:34px">
      <div class="bloc">
        <h3>Référentiel des notions</h3>
        <p>L'ensemble des notions à couvrir, classées par domaine et par catégorie,
           avec leur niveau, leur mode de validation et le type d'exercice visé.</p>
        <button class="bt plein" data-aller="referentiel">Ouvrir le référentiel</button>
      </div>
      <div class="bloc">
        <h3>Exercices du lot A</h3>
        <p>Découverte des composants natifs de Grasshopper pour Rhino 8.
           Chaque exercice comporte un sujet, un corrigé commenté et deux fichiers <code>.gh</code>.</p>
        <button class="bt plein" data-aller="exercices">Parcourir les exercices</button>
      </div>
      <div class="bloc">
        <h3>Comment travailler un exercice</h3>
        <p>Ouvrez le fichier <b>sujet</b> dans Grasshopper, branchez votre résultat
           sur le paramètre <code>REPONSE</code>, puis comparez au corrigé. Dans le fichier
           complet, le corrigé ne s'affiche qu'après avoir basculé l'interrupteur
           <b>AFFICHER LE CORRIGÉ</b>.</p>
      </div>
    </div>
  </section>

  <section id="vue-referentiel" class="vue">
    <div class="barre">
      <input type="search" id="q-ref" placeholder="Rechercher une notion, une catégorie, une description…">
      <select id="f-dom"><option value="">Tous les domaines</option></select>
      <select id="f-niv"><option value="">Tous les niveaux</option></select>
      <span class="compte" id="c-ref"></span>
    </div>
    <div class="cadre-table"><table id="t-ref">
      <thead><tr><th>ID</th><th>Domaine</th><th>Catégorie</th><th>Notion</th>
      <th>Description</th><th>Niveau</th><th>Validation</th><th>Type</th>
      <th>Exercices</th></tr></thead>
      <tbody></tbody></table></div>
  </section>

  <section id="vue-exercices" class="vue">
    <div id="liste-ex">
      <div class="barre">
        <input type="search" id="q-ex" placeholder="Rechercher un exercice…">
        <select id="f-them"><option value="">Toutes les thématiques</option></select>
        <span class="compte" id="c-ex"></span>
      </div>
      <div class="cartes" id="cartes"></div>
    </div>
    <div id="detail-ex" style="display:none"></div>
  </section>
</main>

<script id="donnees" type="application/json">%s</script>
<script>
const D = JSON.parse(document.getElementById('donnees').textContent);
const esc = s => String(s==null?'':s).replace(/[&<>"]/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
/* La racine depend du lot : le lot A et le lot IA ne vivent pas dans le
   meme dossier. On la porte donc par exercice, jamais en global. */
const chemin = (dos, fic, rac) =>
  encodeURI((rac || D.dossier) + '/' + dos + '/' + fic);
const coul = d => D.couleurs[d] || '#94a3b8';

/* Les sous-categories reprennent la teinte de leur domaine, eclaircie : on
   voit d'un coup d'oeil a quel domaine on appartient sans confondre les deux
   niveaux. Melange vers le blanc plutot qu'une opacite, pour rester lisible
   quel que soit le fond. */
function palir(hex, taux){
  const h = (hex || '#94a3b8').replace('#','');
  const r = parseInt(h.slice(0,2),16), v = parseInt(h.slice(2,4),16), b = parseInt(h.slice(4,6),16);
  const m = c => Math.round(c + (255 - c) * taux);
  return 'rgb(' + m(r) + ',' + m(v) + ',' + m(b) + ')';
}

/* ---------- navigation ---------- */
function aller(v){
  document.querySelectorAll('.vue').forEach(s => s.classList.toggle('actif', s.id === 'vue'+'-'+v));
  document.querySelectorAll('nav button').forEach(b => b.classList.toggle('actif', b.dataset.vue === v));
  window.scrollTo(0,0);
  /* Le cadre du tableau n'a de hauteur mesurable qu'une fois sa vue affichee :
     on recalcule apres la bascule, sinon il garde celle de la vue precedente. */
  if (typeof mesurerEntete === 'function') mesurerEntete();
}
/* La ligne d'en-tetes du tableau se colle sous la barre du haut. Cette barre
   change de hauteur quand elle passe sur plusieurs lignes en fenetre etroite :
   une valeur figee ferait flotter les en-tetes au milieu du tableau. On mesure
   donc la hauteur reelle et on la publie en variable CSS. */
function mesurerEntete(){
  const h = document.querySelector('header.top');
  if (!h) return;
  document.documentElement.style.setProperty('--h-entete', h.offsetHeight + 'px');
  /* La barre de filtres se replie differemment selon la largeur : plutot que
     de deviner sa hauteur, on mesure ou commence reellement le cadre. */
  document.querySelectorAll('.cadre-table').forEach(c => {
    c.style.maxHeight = '';
    const haut = c.getBoundingClientRect().top;
    if (haut > 0 && haut < innerHeight){
      c.style.maxHeight = Math.max(240, innerHeight - haut - 18) + 'px';
    }
  });
}
mesurerEntete();
window.addEventListener('resize', mesurerEntete);
if (window.ResizeObserver){
  new ResizeObserver(mesurerEntete).observe(document.querySelector('header.top'));
}

document.querySelectorAll('nav button').forEach(b => b.onclick = () => aller(b.dataset.vue));

/* Le logo ramene a l'accueil ET remet l'acces thematique au premier niveau,
   sinon on retrouverait l'accueil filtre sur la derniere categorie visitee. */
document.getElementById('retour-accueil').onclick = () => {
  fermer();
  NAV = {dom:null, cat:null};
  rendTheme();
  aller('accueil');
};
document.querySelectorAll('[data-aller]').forEach(b => b.onclick = () => aller(b.dataset.aller));

/* ---------- accueil ---------- */
const nbDom = new Set(D.ref.map(r => r['Domaine'])).size;
document.getElementById('chiffres').innerHTML = [
  [D.ref.length, 'notions au référentiel'],
  [nbDom, 'domaines'],
  [D.exos.length, 'exercices'],
  /* On compte les definitions reellement presentes : les multiplier par deux
     supposerait que chaque exercice en a deux, ce qui est faux — IA-07 a pour
     livrable un plugin compile et n'en a aucune. */
  [D.exos.reduce((n,e) => n + (e.dispo.gh_sujet?1:0) + (e.dispo.gh_complet?1:0), 0),
   'définitions Grasshopper']
].map(([n,l]) => `<div class="chiffre"><b>${n}</b><span>${l}</span></div>`).join('');

/* ---------- accès thématique : domaine > catégorie > notions et exercices ---------- */
let NAV = {dom:null, cat:null};
const notionsDe = (dom,cat) => D.ref.filter(r =>
  (!dom || r['Domaine']===dom) && (!cat || r['Catégorie']===cat));
const exosDe = (dom,cat) => D.exos.filter(e =>
  cat ? (e.cats||[]).includes(cat) : (dom ? (e.doms||[]).includes(dom) : true));

function tuile(nom, coulBande, nbN, nbE, act){
  return `<div class="tuile" data-nom="${esc(nom)}" data-act="${act}">
    <div class="bande" style="background:${coulBande}"></div>
    <div><b>${esc(nom)}</b><div class="cpt">${nbN} notion${nbN>1?'s':''}
      · ${nbE} exercice${nbE>1?'s':''}</div></div>
    <div class="fleche">›</div></div>`;
}

function rendTheme(){
  const fil = ['<button data-niv="0">Tous les domaines</button>'];
  if (NAV.dom) fil.push('<span class="sep">›</span>',
    NAV.cat ? `<button data-niv="1">${esc(NAV.dom)}</button>`
            : `<span class="ici">${esc(NAV.dom)}</span>`);
  if (NAV.cat) fil.push('<span class="sep">›</span>', `<span class="ici">${esc(NAV.cat)}</span>`);
  document.getElementById('fil').innerHTML = fil.join('');
  document.querySelectorAll('#fil button').forEach(b => b.onclick = () => {
    if (b.dataset.niv === '0'){ NAV = {dom:null, cat:null}; } else { NAV.cat = null; }
    rendTheme();
  });

  const z = document.getElementById('theme');
  if (!NAV.dom){
    z.innerHTML = '<div class="tuiles">' + D.domaines.map(d =>
      tuile(d, coul(d), notionsDe(d,null).length, exosDe(d,null).length, 'dom')).join('') + '</div>';
  } else if (!NAV.cat){
    const cats = [...new Set(notionsDe(NAV.dom,null).map(r => r['Catégorie']))];
    z.innerHTML = '<div class="tuiles">' + cats.map(c =>
      tuile(c, palir(coul(NAV.dom), 0.45), notionsDe(NAV.dom,c).length,
            exosDe(NAV.dom,c).length, 'cat')).join('') + '</div>';
  } else {
    const ns = notionsDe(NAV.dom, NAV.cat), ex = exosDe(NAV.dom, NAV.cat);
    z.innerHTML = `<div class="duo">
      <div><h3>Notions de cette catégorie (${ns.length})</h3><div class="mini">${
        ns.map(r => `<div class="l"><b>${esc(r['Notion'])}</b>
          <span class="mono">${esc(r['ID'])}</span> — <span>${esc(r['Description'])}</span></div>`).join('')
        || '<div class="l"><span>Aucune notion.</span></div>'}</div></div>
      <div><h3>Exercices rattachés (${ex.length})</h3>${ ex.length
        ? '<div class="cartes">' + ex.map(e => `<div class="carte" data-ex="${e.id}"
            style="border-left-color:${coul(e.dom)}"><div class="cid">${e.id}</div>
            <h4>${esc(e.titre)}</h4><div class="meta"><span class="etiq">${esc(e.niv)}</span>
            <span class="etiq">${e.duree} min</span></div></div>`).join('') + '</div>'
        : '<div class="mini"><div class="l"><span>Aucun exercice ne couvre encore cette catégorie — les lots B, C et G restent à produire.</span></div></div>'}</div>
    </div>`;
    z.querySelectorAll('[data-ex]').forEach(c => c.onclick = () => {
      aller('exercices'); ouvrir(c.dataset.ex);
    });
  }
  z.querySelectorAll('.tuile').forEach(t => t.onclick = () => {
    if (t.dataset.act === 'dom') NAV.dom = t.dataset.nom; else NAV.cat = t.dataset.nom;
    rendTheme();
  });
}
rendTheme();

/* ---------- référentiel ---------- */
const selD = document.getElementById('f-dom'), selN = document.getElementById('f-niv');
D.domaines.forEach(d => selD.add(new Option(d, d)));
[...new Set(D.ref.map(r => r['Niveau']))].filter(Boolean).sort()
  .forEach(n => selN.add(new Option(n, n)));

/* Index notion -> exercices qui la couvrent. Un exercice cite plusieurs
   notions ; on inverse la relation une fois pour toutes. */
const exosParRef = {};
D.exos.forEach(e => (e.ref || '').split(',').forEach(x => {
  const k = x.trim();
  if (k) (exosParRef[k] = exosParRef[k] || []).push(e);
}));

function boutonsExos(idNotion){
  const l = exosParRef[idNotion];
  if (!l || !l.length) return '<span class="vide-cell">—</span>';
  return l.map(e => {
    const chn = e.verdict === 'connaissance';
    return `<button class="bt-exo${chn ? ' charniere' : ''}" data-exo="${e.id}"
      title="${chn ? 'Question charnière' : 'Exercice'} — ${esc(e.titre)}">${e.id}</button>`;
  }).join('');
}

function rendRef(){
  const q = document.getElementById('q-ref').value.toLowerCase();
  const fd = selD.value, fn = selN.value;
  const l = D.ref.filter(r =>
    (!fd || r['Domaine'] === fd) && (!fn || r['Niveau'] === fn) &&
    (!q || [r['ID'],r['Domaine'],r['Catégorie'],r['Notion'],r['Description']]
            .join(' ').toLowerCase().includes(q)));
  document.querySelector('#t-ref tbody').innerHTML = l.map(r => `<tr>
    <td class="mono">${esc(r['ID'])}</td>
    <td><span class="pastille" style="background:${coul(r['Domaine'])}"></span>${esc(r['Domaine'])}</td>
    <td><span class="pastille" style="background:${palir(coul(r['Domaine']),0.45)}"></span>${esc(r['Catégorie'])}</td>
    <td><b>${esc(r['Notion'])}</b></td>
    <td>${esc(r['Description'])}</td><td>${esc(r['Niveau'])}</td>
    <td><span class="etiq">${esc(r['ValidationMode suggéré'])}</span></td>
    <td>${esc(r["Type d'exercice"])}</td>
    <td class="cell-exos">${boutonsExos(r['ID'])}</td></tr>`).join('')
    || '<tr><td colspan="9" class="vide">Aucune notion ne correspond.</td></tr>';
  document.getElementById('c-ref').textContent = l.length + ' / ' + D.ref.length + ' notions';
}
['q-ref','f-dom','f-niv'].forEach(i => document.getElementById(i).oninput = rendRef);

/* Delegation : le tableau est reconstruit a chaque filtre, un gestionnaire pose
   sur chaque bouton serait perdu au premier tri. */
document.querySelector('#t-ref tbody').addEventListener('click', ev => {
  const b = ev.target.closest('.bt-exo');
  if (!b) return;
  aller('exercices');
  ouvrir(b.dataset.exo);
});

rendRef();

/* ---------- exercices ---------- */
const selT = document.getElementById('f-them');
D.themes.forEach(t => selT.add(new Option(t, t)));
function rendEx(){
  const q = document.getElementById('q-ex').value.toLowerCase(), ft = selT.value;
  const l = D.exos.filter(e => (!ft || e.them === ft) &&
    (!q || [e.id,e.titre,e.obj,e.comp,e.enonce].join(' ').toLowerCase().includes(q)));
  document.getElementById('cartes').innerHTML = l.map(e => `
    <div class="carte" data-id="${e.id}" style="border-left-color:${coul(e.dom)}">
      <div class="cid">${e.id}</div><h4>${esc(e.titre)}</h4>
      <div style="color:var(--doux);font-size:13px">${esc(e.them)}</div>
      <div style="color:var(--doux);font-size:12px;margin-top:3px">
        <span class="pastille" style="background:${coul(e.dom)}"></span>${esc(e.dom)}</div>
      <div class="meta"><span class="etiq">${esc(e.niv)}</span>
        <span class="etiq">${e.duree} min</span>
        <span class="etiq">${esc(e.mode)}</span></div>
    </div>`).join('') || '<div class="vide">Aucun exercice ne correspond.</div>';
  document.querySelectorAll('.carte').forEach(c => c.onclick = () => ouvrir(c.dataset.id));
  document.getElementById('c-ex').textContent = l.length + ' / ' + D.exos.length + ' exercices';
}
['q-ex','f-them'].forEach(i => document.getElementById(i).oninput = rendEx);
rendEx();

/* ---------- fiche d'un exercice ---------- */
function ouvrirCharniere(e){
  const dos = e.dossier;
  const lignes = (e.charniere||'').split(String.fromCharCode(10))
    .filter(x => x.trim().length)
    .map(x => x.trim().startsWith('Valeur diagnostique')
      ? `<p class="note"><b>${esc(x)}</b></p>` : `<p>${esc(x)}</p>`).join('');
  document.getElementById('detail-ex').innerHTML = `
   <button class="retour">← Retour à la liste</button>
   <div class="fiche">
    <h2>${e.id} — ${esc(e.titre)}</h2>
    <div class="them">${esc(e.them)}</div>
    <div class="encadre">Cet item <b>n'est pas un exercice noté</b>. Il porte une
      connaissance nécessaire, mais qui s'acquiert et se vérifie par une question,
      non par un montage : la construire dans Grasshopper mesurerait la mémoire,
      pas la compétence.</div>
    <table class="infos">
      <tr><td>Référence au référentiel</td><td class="mono">${esc(e.ref)}</td></tr>
      <tr><td>Nature</td><td>Connaissance — question charnière</td></tr>
      ${e.bloom ? `<tr><td>Case Bloom (révisée)</td><td>${esc(e.bloom)}</td></tr>` : ''}
      <tr><td>Niveau</td><td>${esc(e.niv)}</td></tr>
      <tr><td>Mode de validation</td><td>— (non notée)</td></tr>
    </table>
    ${e.contexte ? `<h4 class="sous">Contexte</h4><p>${esc(e.contexte)}</p>` : ''}
    <h3 class="sect">LA QUESTION</h3>
    ${lignes}
    <h3 class="sect">COMMENT L'EMPLOYER</h3>
    <ul class="puces">
      <li>Avant l'exercice qui mobilise cette connaissance, pas après : elle en est un prérequis.</li>
      <li>Poser la question à main levée, relever la répartition des réponses,
          et n'expliquer que si une réponse fausse est majoritaire.</li>
      <li>La valeur est dans la mauvaise réponse : elle nomme la représentation à corriger.</li>
    </ul>
    <h4 class="sous">Énoncé d'origine, conservé pour mémoire</h4>
    <p class="note">${esc(e.enonce_origine || e.enonce)}</p>
    ${e.dispo.images ? `<h3 class="sect">DÉMONSTRATION FACULTATIVE</h3>` : ''}
    <p class="note">Le fichier ${e.id}_complet.gh reste disponible comme support de
      démonstration au vidéoprojecteur. Il n'est pas à faire construire.</p>
    <figure><img loading="lazy" src="${chemin(dos, 'Illustrations/web/'+e.id+'_canvas_corrige.jpg', e.racine)}" alt="">
      <figcaption>Le montage de démonstration —
        <a href="${chemin(dos, 'Illustrations/'+e.id+'_canvas_corrige.png', e.racine)}" target="_blank">pleine résolution</a></figcaption></figure>
    <div class="telech">
      ${e.dispo.pdf ? `<a class="bt plein" href="${chemin(dos, e.id+'_fiche.pdf', e.racine)}" download>⬇ Fiche PDF</a>` : ''}
      ${e.dispo.docx ? `<a class="bt" href="${chemin(dos, e.id+'_fiche.docx', e.racine)}" download>⬇ Fiche Word</a>` : ''}
      ${e.dispo.gh_complet ? `<a class="bt" href="${chemin(dos, e.id+'_complet.gh', e.racine)}" download>⬇ Définition ${e.id}_complet.gh</a>` : ''}
      <button class="bt" onclick="window.print()">🖨 Imprimer / PDF</button>
    </div>
   </div>`;
  document.getElementById('liste-ex').style.display = 'none';
  document.getElementById('detail-ex').style.display = 'block';
  document.querySelector('.retour').onclick = fermer;
  window.scrollTo(0,0);
}
function ouvrir(id){
  const e = D.exos.find(x => x.id === id);
  if (e.verdict === 'connaissance'){ ouvrirCharniere(e); return; }
  const dos = e.dossier;
  const li = a => a.map(x => `<li>${esc(x)}</li>`).join('');
  document.getElementById('detail-ex').innerHTML = `
   <button class="retour">← Retour à la liste</button>
   <div class="fiche">
    <h2>${e.id} — ${esc(e.titre)}</h2>
    <div class="them">${esc(e.them)}</div>
    <table class="infos">
      <tr><td>Référence au référentiel</td><td class="mono">${esc(e.ref)}</td></tr>
      ${e.competence && e.competence !== '—' ? `<tr><td>Compétence visée</td><td>${esc(e.competence)}</td></tr>` : ''}
      ${e.bloom ? `<tr><td>Case Bloom (révisée)</td><td>${esc(e.bloom)}</td></tr>` : ''}
      <tr><td>Niveau</td><td>${esc(e.niv)}</td></tr>
      <tr><td>Durée cible</td><td>${e.duree} minutes</td></tr>
      <tr><td>Prérequis</td><td>${esc(e.prereq)}</td></tr>
      ${e.verdict === 'connaissance'
        ? `<tr><td>Mode de validation</td><td>— (question charnière, non notée)</td></tr>`
        : `<tr><td>Mode de validation</td><td>${esc(e.mode)} — tolérance ${esc(e.tol)}</td></tr>
           <tr><td>Solution de référence</td><td>${esc(e.nb)} composants</td></tr>`}
      <tr><td>Gamification associée</td><td>${esc(e.gamif)}</td></tr>
    </table>

    <h3 class="sect">SUJET</h3>
    <h4 class="sous">Compétence visée</h4><p>${esc(e.obj)}</p>
    ${e.contexte ? `<h4 class="sous">Contexte</h4><p>${esc(e.contexte)}</p>` : ''}
    <h4 class="sous">Énoncé</h4><div class="encadre">${esc(e.enonce)}</div>
    ${e.dispo.images ? `<figure><img loading="lazy" src="${chemin(dos, 'Illustrations/web/'+e.id+'_canvas_sujet.jpg', e.racine)}" alt="">
      <figcaption>Le canvas à l'ouverture de ${e.id}_sujet.gh —
        <a href="${chemin(dos, 'Illustrations/'+e.id+'_canvas_sujet.png', e.racine)}" target="_blank">pleine résolution</a></figcaption></figure>` : ''}
    <h4 class="sous">Ce qui vous est fourni</h4><p>${esc(e.depart)}</p>
    <h4 class="sous">Ce qui est attendu</h4><p>${esc(e.att)}</p>
    <p class="note">La consigne ne nomme aucun composant, et c'est délibéré :
      nommer l'outil reviendrait à donner la réponse. La liste des composants
      figure avec le corrigé.</p>
    <h4 class="sous">Barème</h4><p>${esc(e.bareme)}</p>

    <h3 class="sect">CORRIGÉ</h3>
    <label class="interrupteur"><input type="checkbox" id="sw-corrige">
      <span><b>Afficher le corrigé</b> — comme dans la définition Grasshopper,
      il reste masqué tant que vous ne l'avez pas demandé.</span></label>
    <div id="corrige" class="masque">
      ${e.dispo.images ? `<figure><img loading="lazy" src="${chemin(dos, 'Illustrations/web/'+e.id+'_canvas_corrige.jpg', e.racine)}" alt="">
        <figcaption>La zone corrigé, autonome : aucun câble ne la relie à la zone sujet —
          <a href="${chemin(dos, 'Illustrations/'+e.id+'_canvas_corrige.png', e.racine)}" target="_blank">pleine résolution</a></figcaption></figure>` : ''}
      <h4 class="sous">Marche à suivre</h4><ol class="etapes">${li(e.etapes)}</ol>
      ${e.erreur ? `<h4 class="sous">L'erreur attendue</h4>
        <p class="note">Elle est diagnostique : elle dit ce que l'apprenant a mal
        compris, là où un simple « faux » ne dirait rien.</p>
        <div class="encadre rouge">${esc(e.erreur)}</div>` : ''}
      <h4 class="sous">Pièges fréquents</h4><ul class="puces">${li(e.pieges)}</ul>
      <h4 class="sous">Composants de la solution de référence</h4><p>${esc(e.comp)}</p>
      ${e.donnees_note ? `<h4 class="sous">Pourquoi ce jeu de données</h4><p>${esc(e.donnees_note)}</p>` : ''}
      ${e.limite ? `<h4 class="sous">Limite de la correction automatique</h4><div class="encadre">${esc(e.limite)}</div>` : ''}
      ${e.alerte ? `<h4 class="sous">Note au formateur</h4><div class="encadre">${esc(e.alerte)}</div>` : ''}
      ${e.manuel.length ? '<h4 class="sous">Réglages à poser à la main</h4><div class="encadre rouge"><ul class="puces">'+li(e.manuel)+'</ul></div>' : ''}
      <h4 class="sous">Pour aller plus loin</h4><ul class="puces">${li(e.var)}</ul>
    </div>

    <div class="telech">
      ${e.dispo.pdf ? `<a class="bt plein" href="${chemin(dos, e.id+'_fiche.pdf', e.racine)}" download>⬇ Fiche PDF</a>` : ''}
      ${e.dispo.docx ? `<a class="bt" href="${chemin(dos, e.id+'_fiche.docx', e.racine)}" download>⬇ Fiche Word</a>` : ''}
      ${e.dispo.docx_sujet ? `<a class="bt" href="${chemin(dos, e.id+'_fiche_sujet.docx', e.racine)}" download>⬇ Fiche Word — sujet seul</a>` : ''}
      ${e.dispo.gh_sujet ? `<a class="bt" href="${chemin(dos, e.id+'_sujet.gh', e.racine)}" download>⬇ Définition ${e.id}_sujet.gh</a>` : ''}
      ${e.dispo.gh_complet ? `<a class="bt" href="${chemin(dos, e.id+'_complet.gh', e.racine)}" download>⬇ Définition ${e.id}_complet.gh</a>` : ''}
      <button class="bt" onclick="window.print()">🖨 Imprimer / PDF</button>
    </div>
    ${!e.dispo.pdf && !e.dispo.gh_sujet ? `<p class="note">Les livrables de cet exercice — définition Grasshopper,
      fiches Word et PDF — restent à produire. La fiche ci-dessus fait foi en attendant.</p>` : ''}
   </div>`;
  document.getElementById('liste-ex').style.display = 'none';
  document.getElementById('detail-ex').style.display = 'block';
  document.querySelector('.retour').onclick = fermer;
  document.getElementById('sw-corrige').onchange = ev =>
    document.getElementById('corrige').classList.toggle('masque', !ev.target.checked);
  window.scrollTo(0,0);
}
function fermer(){
  document.getElementById('detail-ex').style.display = 'none';
  document.getElementById('liste-ex').style.display = 'block';
  window.scrollTo(0,0);
}
</script></body></html>
""" % (favicon,
       CSS + (PORTAIL_CSS if PROTEGE else u""),
       logo_html, VERSION, INDICE, DATE, js)

    if PROTEGE:
        html = html.replace(
            u"</script></body></html>",
            (PORTAIL % (json.dumps(PROTEGE[0]), json.dumps(PROTEGE[1]),
                        json.dumps(PROTEGE[2])))
            + u"</script></body></html>")

    io.open(SORTIE, "w", encoding="utf-8").write(html)
    io.open(SORTIE_INDEX, "w", encoding="utf-8").write(html)
    print(u"Notions : %d | Exercices : %d" % (len(ref), len(exos)))
    print(u"écrit : %s  (%.0f Ko)" % (SORTIE, os.path.getsize(SORTIE) / 1024.0))


if __name__ == "__main__":
    main()
