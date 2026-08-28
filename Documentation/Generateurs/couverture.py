# -*- coding: utf-8 -*-
"""Etat de couverture du referentiel par les exercices produits.

Rend, domaine par domaine et categorie par categorie, le nombre de notions et
le nombre d'exercices qui les mobilisent. Sert a savoir ou porter l'effort
plutot qu'a le deviner.
"""
import io
import os
import re
import sys

ICI = os.path.dirname(os.path.abspath(__file__))
PROJET = os.path.abspath(os.path.join(ICI, "..", ".."))
if ICI not in sys.path:
    sys.path.insert(0, ICI)

CLASSEUR = os.path.join(PROJET, "Fondamentaux Grasshopper - IndB - 26-08-2026.xlsx")


def notions():
    import openpyxl
    w = openpyxl.load_workbook(CLASSEUR, read_only=True, data_only=True)
    ws = w[u"Référentiel"]
    ent = [c for c in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
    idx = dict((n, i) for i, n in enumerate(ent) if n)
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[idx[u"ID"]]:
            continue
        out.append({
            u"id": r[idx[u"ID"]],
            u"dom": r[idx[u"Domaine"]] or u"",
            u"cat": r[idx[u"Catégorie"]] or u"",
            u"notion": r[idx[u"Notion"]] or u"",
            u"niv": r[idx[u"Niveau"]] or u"",
        })
    return out


def exercices():
    """Tous les exercices, tous lots confondus (registre unique)."""
    from lots import TOUS
    return TOUS


def rang(d):
    n = u""
    for c in d:
        if c.isdigit():
            n += c
        else:
            break
    return (int(n) if n else 999, d)


def main():
    ns = notions()
    ex = exercices()

    par_ref = {}
    for e in ex:
        for rid in re.findall(u"REF-[0-9]+", e.get("ref", u"") or u""):
            par_ref.setdefault(rid, []).append(e["id"])

    # regroupement domaine > categorie
    doms = {}
    for n in ns:
        doms.setdefault(n[u"dom"], {}).setdefault(n[u"cat"], []).append(n)

    lignes = []
    tot_n = tot_c = 0
    vides = []
    lignes.append(u"# Couverture du référentiel par les exercices\n")
    lignes.append(u"%d notions, %d exercices produits.\n" % (len(ns), len(ex)))

    for dom in sorted(doms, key=rang):
        cats = doms[dom]
        nn = sum(len(v) for v in cats.values())
        ee = set()
        for v in cats.values():
            for n in v:
                ee |= set(par_ref.get(n[u"id"], []))
        couv = sum(1 for v in cats.values() for n in v if par_ref.get(n[u"id"]))
        tot_n += nn
        tot_c += couv
        lignes.append(u"\n## %s" % dom)
        lignes.append(u"\n%d notions, %d couvertes, %d exercices.\n"
                      % (nn, couv, len(ee)))
        lignes.append(u"| Catégorie | Notions | Couvertes | Exercices |")
        lignes.append(u"|---|---:|---:|---|")
        for cat in sorted(cats):
            v = cats[cat]
            e2 = set()
            c2 = 0
            for n in v:
                l = par_ref.get(n[u"id"], [])
                e2 |= set(l)
                if l:
                    c2 += 1
            if c2 == 0:
                vides.append((dom, cat, len(v)))
            lignes.append(u"| %s | %d | %d | %s |"
                          % (cat, len(v), c2,
                             u", ".join(sorted(e2)) if e2 else u"—"))

    lignes.append(u"\n---\n")
    lignes.append(u"## Catégories sans aucun exercice\n")
    lignes.append(u"**%d catégories**, soit **%d notions** non couvertes.\n"
                  % (len(vides), sum(v[2] for v in vides)))
    lignes.append(u"| Domaine | Catégorie | Notions |")
    lignes.append(u"|---|---|---:|")
    for d, c, k in sorted(vides, key=lambda x: (rang(x[0]), x[1])):
        lignes.append(u"| %s | %s | %d |" % (d, c, k))

    lignes.append(u"\n**Total : %d notions sur %d couvertes (%.0f %%).**"
                  % (tot_c, tot_n, 100.0 * tot_c / tot_n))

    dest = os.path.join(ICI, u"COUVERTURE.md")
    io.open(dest, "w", encoding="utf-8").write(u"\n".join(lignes))

    print(u"notions : %d | couvertes : %d (%.0f %%)"
          % (tot_n, tot_c, 100.0 * tot_c / tot_n))
    print(u"catégories sans exercice : %d, soit %d notions"
          % (len(vides), sum(v[2] for v in vides)))
    print(u"écrit : %s" % dest)
    return vides


if __name__ == "__main__":
    main()
